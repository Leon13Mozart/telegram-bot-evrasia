import asyncio
import calendar
import csv
import html
import hashlib
import hmac
import json
import time
from urllib.parse import parse_qsl
import logging
import os
import re
import secrets
import sqlite3
import threading
import httpx
from io import StringIO
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from telegram import (
    BotCommand,
    BotCommandScopeAllGroupChats,
    BotCommandScopeAllPrivateChats,
    BotCommandScopeChat,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
    WebAppInfo,
)
from telegram.constants import ChatMemberStatus

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

from telegram.ext import (
    Application,
    CallbackQueryHandler,
    ChatMemberHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)


# ============================================================
# НАЛАШТУВАННЯ
# ============================================================

TOKEN = "8835839482:AAFQ0yWwNdZ7dHUzJKPernaOVo_HMUNL24g"

ADMINS = [
    929200380,
    395523040,
]

DEVELOPER_ID = 929200380
DEFAULT_MANAGER_ID = 395523040

DATABASE = os.getenv("DATABASE_PATH", "telegram_stats.db")

WEB_APP_URL = os.getenv(
    "WEB_APP_URL",
    "https://api.185-237-96-40.sslip.io"
).rstrip("/")

WEB_ALLOWED_ORIGINS = [
    origin.strip().rstrip("/")
    for origin in os.getenv(
        "WEB_ALLOWED_ORIGINS",
        "https://api.185-237-96-40.sslip.io,https://evrasia-masterclass.web.app,https://evrasia-masterclass.firebaseapp.com",
    ).split(",")
    if origin.strip()
]

WEB_DEBUG_USER_ID = os.getenv("WEB_DEBUG_USER_ID", "").strip()

TIMEZONE = ZoneInfo("Europe/Kyiv")

AUTO_REPORT_CHAT_IDS = ADMINS

REPORT_HOUR = 23
REPORT_MINUTE = 59


# Google Apps Script, куди бот записує нові реєстрації.
GOOGLE_MK_WEB_APP_URL = os.getenv(
    "GOOGLE_MK_WEB_APP_URL",
    "https://script.google.com/macros/s/AKfycbzHFmihuVlvb9fGDplo0Yxdw6-00y-UztV-VLMO/exec",
).strip()

# Існуюча Google-таблиця реєстрацій. Нову таблицю бот НЕ створює.
GOOGLE_MK_SHEET_ID = os.getenv(
    "GOOGLE_MK_SHEET_ID",
    "1wy1a3aZ6w5B56R9vbWxuIq6qFcFzOiiOpMczF5frfJM",
).strip()

GOOGLE_MK_REGISTRATIONS_SHEET = os.getenv(
    "GOOGLE_MK_REGISTRATIONS_SHEET",
    "Всі реєстрації",
).strip()

# Не звертаємося до Google при кожному кліку на сайті.
GOOGLE_MK_READ_CACHE_SECONDS = int(
    os.getenv("GOOGLE_MK_READ_CACHE_SECONDS", "30")
)

_google_mk_read_lock = threading.RLock()
_google_mk_last_read_at = 0.0


# Відповідність назв Telegram-груп назвам ресторанів у Google-таблиці.
# Пошук виконується за фрагментом назви групи, без урахування регістру.
GOOGLE_MK_RESTAURANT_ALIASES = {
    "retroville": "пр-т Правди, 47",
    "маяковського 5в": "пр-т Червоної калини, 5в",
    "маяковського 44а": "пр-т Червоної калини, 44а",
    "оболонський": "пр-т Оболонський, 19",
    "григоренко": "Григоренко пр-т, 22/20",
    "драгоманова": "Драгоманова, 40г",
    "нивках": "пр-т Берестейський, 84в",
    "рудницького": "Рудницького, 5",
    "святошинська": "Вишневе, Святошинська, 27е",
    "палац спорту": "Рогніденська, 5/14 (Шота)",
    "лівобереж": "Раїси Окіпної, 3",
    "республіка": "Кільцева Дорога, 1 (Respublika park)",
}

logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(message)s",
    level=logging.INFO,
)

# Тимчасові дані для створення та підтвердження розсилки.
broadcast_wait = {}
broadcast_message = {}
delete_message_cache = {}

# Тимчасовий стан приватного майстра налаштування відповідей МК.
mk_config_state = {}
mk_name_wait = {}


# ============================================================
# РОБОТА З БАЗОЮ ДАНИХ
# ============================================================


def register_group(chat):
    with sqlite3.connect(DATABASE) as conn:
        conn.execute('\n            INSERT OR IGNORE INTO groups(\n                group_id,\n                title\n            )\n            VALUES(?,?)\n            ', (chat.id, chat.title))
        conn.commit()


def save_event(chat, user, action):
    register_group(chat)
    username = user.username or ''
    with sqlite3.connect(DATABASE) as conn:
        conn.execute('\n            INSERT INTO events(\n                group_id,\n                user_id,\n                username,\n                first_name,\n                action,\n                event_time\n            )\n            VALUES(?,?,?,?,?,?)\n            ', (chat.id, user.id, username, user.first_name, action, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()


def check_db():
    with sqlite3.connect(DATABASE) as conn:
        tables = conn.execute("\n            SELECT name \n            FROM sqlite_master\n            WHERE type='table'\n            ").fetchall()
    print('Таблиці SQLite:', tables)


def init_db():
    with sqlite3.connect(DATABASE) as conn:
        conn.execute('\n        CREATE TABLE IF NOT EXISTS groups(\n            group_id INTEGER PRIMARY KEY,\n            title TEXT\n        )\n        ')
        conn.execute('\n        CREATE TABLE IF NOT EXISTS events(\n            id INTEGER PRIMARY KEY AUTOINCREMENT,\n            group_id INTEGER,\n            user_id INTEGER,\n            username TEXT,\n            first_name TEXT,\n            action TEXT,\n            event_time TEXT\n        )\n        ')
        conn.execute('\n        CREATE TABLE IF NOT EXISTS last_broadcast(\n            id INTEGER PRIMARY KEY AUTOINCREMENT,\n            group_id INTEGER NOT NULL,\n            message_id INTEGER NOT NULL\n        )\n        ')
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS mk_settings(
                group_id INTEGER PRIMARY KEY,
                response_text TEXT NOT NULL,
                confirmation_text TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS mk_pending(
                token TEXT PRIMARY KEY,
                group_id INTEGER NOT NULL,
                group_title TEXT,
                requester_id INTEGER NOT NULL,
                requester_name TEXT,
                child_name TEXT NOT NULL,
                source_message_id INTEGER,
                created_at TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'waiting_start'
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS mk_user_state(
                user_id INTEGER PRIMARY KEY,
                token TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS mk_contacts(
                user_id INTEGER PRIMARY KEY,
                phone_number TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS mk_registrations(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                group_id INTEGER NOT NULL,
                group_title TEXT,
                requester_id INTEGER NOT NULL,
                requester_name TEXT,
                child_name TEXT NOT NULL,
                phone_number TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.commit()
    print('✅ База данных готова')


# ============================================================
# ОТРИМАННЯ СТАТИСТИКИ
# ============================================================


def count_events(group_id, action=None, days=None):
    sql = '\n    SELECT COUNT(*)\n    FROM events\n    WHERE group_id=?\n    '
    params = [group_id]
    if action:
        sql += ' AND action=?'
        params.append(action)
    if days is not None:
        date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d %H:%M:%S')
        sql += ' AND event_time>=?'
        params.append(date)
    with sqlite3.connect(DATABASE) as conn:
        result = conn.execute(sql, tuple(params)).fetchone()
        return result[0] if result else 0


def total_stats(group_id):
    joins = count_events(group_id, 'join')
    leaves = count_events(group_id, 'leave')
    return (joins, leaves)


def today_stats(group_id):
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    return count_events_from_date(group_id, today)


def week_stats(group_id):
    now = datetime.now()
    monday = now - timedelta(days=now.weekday())
    monday = monday.replace(hour=0, minute=0, second=0, microsecond=0)
    return count_events_from_date(group_id, monday)


def month_stats(group_id):
    now = datetime.now()
    first_day = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return count_events_from_date(group_id, first_day)


def count_events_from_date(group_id, date):
    joins = 0
    leaves = 0
    sql = '\n    SELECT action, COUNT(*)\n    FROM events\n    WHERE group_id=?\n      AND event_time>=?\n    GROUP BY action\n    '
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(sql, (group_id, date.strftime('%Y-%m-%d %H:%M:%S'))).fetchall()
    for action, count in rows:
        if action == 'join':
            joins = count
        elif action == 'leave':
            leaves = count
    return (joins, leaves)


def get_groups():
    with sqlite3.connect(DATABASE) as conn:
        cursor = conn.execute('\n            SELECT group_id, title\n            FROM groups\n            ORDER BY title\n            ')
        return cursor.fetchall()


# ============================================================
# ПЕРЕВІРКА ДОСТУПУ
# ============================================================


def is_admin(user_id):
    return user_id in ADMINS


# ============================================================
# ОБРОБНИК ВХОДУ / ВИХОДУ УЧАСНИКІВ
# ============================================================


async def chat_member_update(update: Update, context: ContextTypes.DEFAULT_TYPE):
    result = update.chat_member
    old_status = result.old_chat_member.status
    new_status = result.new_chat_member.status
    chat = result.chat
    user = result.new_chat_member.user
    joined = old_status in (ChatMemberStatus.LEFT, ChatMemberStatus.BANNED) and new_status in (ChatMemberStatus.MEMBER, ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.OWNER)
    left = old_status in (ChatMemberStatus.MEMBER, ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.OWNER) and new_status in (ChatMemberStatus.LEFT, ChatMemberStatus.BANNED)
    if joined:
        save_event(chat, user, 'join')
        logging.info(f'{user.full_name} приєднався до {chat.title}')
    elif left:
        save_event(chat, user, 'leave')
        logging.info(f'{user.full_name} покинув {chat.title}')


# ============================================================
# КОМАНДИ БОТА
# ============================================================


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Посилання для завершення запису на майстер-клас доступне всім користувачам.
    if context.args and context.args[0].startswith("mk_"):
        token = context.args[0][3:]
        await begin_mk_private_registration(update, context, token)
        return

    if not is_admin(update.effective_user.id):
        return

    text = (
        "📊 Бот статистики працює.\n\n"
        "Доступні команди:\n\n"
        "/groups — список груп\n"
        "/stats — статистика\n"
        "/addgroup — додати поточну групу\n"
        "/excel — експорт Excel\n"
        "Автозвіт — останнього дня місяця о 23:59\n"
        "/broadcast — розсилка\n"
        "/deletebroadcast — видалити останню розсилку\n"
        "/delete_id — видалити повідомлення за ID\n\n"
        "Запис на МК:\n"
        "/setmkresponse ТЕКСТ — налаштувати відповідь у групі\n"
        "/setmkconfirm ТЕКСТ — налаштувати підтвердження\n"
        "/mklist — останні записи"
    )
    await update.message.reply_text(text)


async def addgroup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMINS:
        return
    chat = update.effective_chat
    if chat.type not in ('group', 'supergroup'):
        await update.message.reply_text('Цю команду можна використовувати лише в групі.')
        return
    register_group(chat)
    await update.message.reply_text(f'✅ Групу збережено\n\nНазва: {chat.title}\nID: {chat.id}')


def count_events_between(group_id, action, start_date, end_date):
    with sqlite3.connect(DATABASE) as conn:
        result = conn.execute(
            """
            SELECT COUNT(*)
            FROM events
            WHERE group_id=?
              AND action=?
              AND event_time>=?
              AND event_time<?
            """,
            (
                group_id,
                action,
                start_date.strftime("%Y-%m-%d %H:%M:%S"),
                end_date.strftime("%Y-%m-%d %H:%M:%S"),
            ),
        ).fetchone()
    return result[0] if result else 0


def create_monthly_excel(report_date):
    groups = get_groups()
    if not groups:
        return None

    months = {
        1: "Січень",
        2: "Лютий",
        3: "Березень",
        4: "Квітень",
        5: "Травень",
        6: "Червень",
        7: "Липень",
        8: "Серпень",
        9: "Вересень",
        10: "Жовтень",
        11: "Листопад",
        12: "Грудень",
    }

    month_start_dt = report_date.replace(
        day=1,
        hour=0,
        minute=0,
        second=0,
        microsecond=0,
    )

    if report_date.month == 12:
        next_month_dt = month_start_dt.replace(
            year=report_date.year + 1,
            month=1,
        )
    else:
        next_month_dt = month_start_dt.replace(
            month=report_date.month + 1,
        )

    month_name = months[report_date.month]
    month_start = month_start_dt.strftime("%d.%m.%Y")
    month_end = report_date.strftime("%d.%m.%Y")

    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика"

    ws["A1"] = "📊 ЗВІТ ПО ГРУПАХ"
    ws["A2"] = f"{month_name} {report_date.year}: {month_start} — {month_end}"
    ws["A3"] = f"Дата формування: {report_date.strftime('%d.%m.%Y %H:%M')}"

    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"].font = Font(bold=True)
    ws["A3"].font = Font(italic=True)

    row = 6

    for group_id, title in groups:
        month_join = count_events_between(
            group_id,
            "join",
            month_start_dt,
            next_month_dt,
        )
        month_leave = count_events_between(
            group_id,
            "leave",
            month_start_dt,
            next_month_dt,
        )
        growth = month_join - month_leave

        ws.cell(row=row, column=1).value = f"🍣 {title}"
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1

        data = [
            ("📅 Період", f"{month_start} — {month_end}"),
            ("➕ Додалося", month_join),
            ("➖ Вийшло", month_leave),
            ("📈 Приріст", growth),
        ]

        for key, value in data:
            ws.cell(row=row, column=1).value = key
            ws.cell(row=row, column=2).value = value
            row += 1

        row += 2

    for column in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = length + 4

    filename = f"Статистика_{report_date.strftime('%d-%m-%Y')}.xlsx"
    filepath = Path(filename)
    wb.save(filepath)
    return filepath


async def excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMINS:
        return

    filepath = create_monthly_excel(datetime.now(TIMEZONE))

    if filepath is None:
        await update.message.reply_text("Немає збережених груп.")
        return

    try:
        with filepath.open("rb") as file:
            await update.message.reply_document(
                document=file,
                filename=filepath.name,
                caption="📊 Звіт по всіх групах",
            )
    finally:
        filepath.unlink(missing_ok=True)


def next_month_report_time(now):
    last_day = calendar.monthrange(now.year, now.month)[1]
    target = now.replace(
        day=last_day,
        hour=REPORT_HOUR,
        minute=REPORT_MINUTE,
        second=0,
        microsecond=0,
    )

    if target <= now:
        if now.month == 12:
            year = now.year + 1
            month = 1
        else:
            year = now.year
            month = now.month + 1

        last_day = calendar.monthrange(year, month)[1]
        target = datetime(
            year,
            month,
            last_day,
            REPORT_HOUR,
            REPORT_MINUTE,
            tzinfo=TIMEZONE,
        )

    return target


async def send_automatic_monthly_report(application):
    report_date = datetime.now(TIMEZONE)
    filepath = create_monthly_excel(report_date)

    if filepath is None:
        logging.warning("Автоматичний звіт не створено: немає груп.")
        return

    try:
        for chat_id in AUTO_REPORT_CHAT_IDS:
            try:
                with filepath.open("rb") as file:
                    await application.bot.send_document(
                        chat_id=chat_id,
                        document=file,
                        filename=filepath.name,
                        caption=(
                            "📊 Автоматичний місячний звіт\n"
                            f"Період: {report_date.replace(day=1).strftime('%d.%m.%Y')} — "
                            f"{report_date.strftime('%d.%m.%Y')}"
                        ),
                    )
            except Exception:
                logging.exception(
                    "Не вдалося надіслати місячний звіт адміністратору %s",
                    chat_id,
                )
    finally:
        filepath.unlink(missing_ok=True)


async def monthly_report_scheduler(application):
    while True:
        now = datetime.now(TIMEZONE)
        target = next_month_report_time(now)
        delay = max((target - now).total_seconds(), 1)

        logging.info(
            "Наступний автоматичний Excel-звіт: %s",
            target.strftime("%d.%m.%Y %H:%M %Z"),
        )

        await asyncio.sleep(delay)
        await send_automatic_monthly_report(application)


async def post_init(application):
    # Швидкий список команд у меню Telegram для особистого чату.
    private_commands = [
        BotCommand("start", "Відкрити меню бота"),
        BotCommand("groups", "Показати список груп"),
        BotCommand("stats", "Показати статистику груп"),
        BotCommand("excel", "Отримати Excel-звіт"),
        BotCommand("broadcast", "Створити розсилку"),
        BotCommand("deletebroadcast", "Видалити останню розсилку"),
        BotCommand("delete_id", "Видалити повідомлення за ID"),
        BotCommand("setmkresponse", "Налаштувати відповідь для МК"),
        BotCommand("setmkconfirm", "Налаштувати підтвердження МК"),
        BotCommand("mklist", "Показати останні записи на МК"),
        BotCommand("cancelmk", "Скасувати налаштування МК"),
    ]

    # У групах залишаються тільки команди, які не розкривають адмін-налаштування.
    group_commands = [
        BotCommand("addgroup", "Додати поточну групу"),
    ]

    await application.bot.set_my_commands(
        private_commands,
        scope=BotCommandScopeAllPrivateChats(),
    )
    await application.bot.set_my_commands(
        group_commands,
        scope=BotCommandScopeAllGroupChats(),
    )

    application.create_task(
        monthly_report_scheduler(application),
        name="monthly-report-scheduler",
    )


async def groups(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user

    if not user or not can_manage_mk(user.id):
        return

    groups_list = get_accessible_groups(user.id)

    if not groups_list:
        await update.message.reply_text(
            "Немає доступних груп."
        )
        return

    text = "📋 Список груп:\n\n"

    for group_id, title in groups_list:

        text += (
            f"🏷 {title}\n"
            f"{group_id}\n\n"
        )

    await update.message.reply_text(text)

async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMINS:
        return
    groups = get_groups()
    if not groups:
        await update.message.reply_text('Немає збережених груп.')
        return
    now = datetime.now()
    today_date = now.strftime('%d.%m.%Y')
    week_start = (now - timedelta(days=now.weekday())).strftime('%d.%m.%Y')
    week_end = now.strftime('%d.%m.%Y')
    month_start = now.replace(day=1).strftime('%d.%m.%Y')
    month_end = now.strftime('%d.%m.%Y')
    months = {1: 'Січень', 2: 'Лютий', 3: 'Березень', 4: 'Квітень', 5: 'Травень', 6: 'Червень', 7: 'Липень', 8: 'Серпень', 9: 'Вересень', 10: 'Жовтень', 11: 'Листопад', 12: 'Грудень'}
    month_name = months[now.month]
    text = '📊 <b>Статистика груп</b>\n\n'
    for group_id, title in groups:
        today_join, today_leave = today_stats(group_id)
        week_join, week_leave = week_stats(group_id)
        month_join, month_leave = month_stats(group_id)
        total_join = count_events(group_id, 'join')
        total_leave = count_events(group_id, 'leave')
        try:
            members = await context.bot.get_chat_member_count(group_id)
        except Exception:
            members = 'Невідомо'
        growth = total_join - total_leave
        if isinstance(growth, int):
            growth_text = f'+{growth}' if growth >= 0 else str(growth)
        else:
            growth_text = 'Невідомо'
        text += f'🍣 <b>{title}</b>\n\n📅 <b>{month_name} {now.year}</b>\n🗓 {month_start} — {month_end}\n➕ Додалося: {month_join}\n➖ Вийшло: {month_leave}\n\n👥 <b>Учасників:</b> {members}\n📈 <b>Приріст:</b> {growth_text}\n\n📊 <b>Загалом</b>\n➕ Додалося: {total_join}\n➖ Вийшло: {total_leave}\n\n────────────────────\n\n'
    await update.message.reply_text(text, parse_mode='HTML')


async def ping(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    await update.message.reply_text('🏓 Pong\nБот працює.')



# ============================================================
# ЗАПИС НА МАЙСТЕР-КЛАСИ
# ============================================================

DEFAULT_MK_RESPONSE = (
    "Натисніть кнопку нижче та надішліть номер телефону боту в особистому чаті."
)
DEFAULT_MK_CONFIRMATION = "✅ Запис для «{name}» підтверджено."

MK_TRIGGER_RE = re.compile(
    r"(?:"
    # Українська
    r"\bзаписат(?:и|ись|ися)\b|"
    r"\bзапиш(?:іть|и|емо|ете)\b|"
    r"\bзаписуєт(?:е|ься)\b|"
    r"\bзаписуйт(?:е|есь)\b|"
    r"\bхоч(?:у|емо|емося|еться)\s+(?:записат(?:и|ись|ися)|на\s+майстер[-\s]?клас)\b|"
    r"\bможна\s+(?:мене\s+|нас\s+|дитину\s+|дитину\s+мою\s+)?"
    r"(?:записат(?:и|ись|ися)|записатися)\b|"
    r"\bпотрібно\s+записат(?:и|ись|ися)\b|"
    r"\bпрошу\s+записат(?:и|ись|ися)\b|"
    r"\bхотіл(?:а|и|о)\s+б\s+записат(?:и|ись|ися)\b|"
    r"\bчи\s+можна\s+(?:сьогодні\s+|на\s+сьогодні\s+|на\s+\d{1,2}[./]\d{1,2}\s+)?записат(?:ись|ися|и)\b|"
    # Російська
    r"\bзаписат(?:ь|ься)\b|"
    r"\bзапиш(?:ите|и|ем|ете)\b|"
    r"\bзаписыва(?:йте|йтесь|ете)\b|"
    r"\bхоч(?:у|им|ется)\s+(?:записат(?:ь|ься)|на\s+мастер[-\s]?класс)\b|"
    r"\bможно\s+(?:меня\s+|нас\s+|ребенка\s+|ребёнка\s+)?"
    r"(?:записат(?:ь|ься))\b|"
    r"\bнужно\s+записат(?:ь|ься)\b|"
    r"\bпрошу\s+записат(?:ь|ься)\b|"
    r"\bхотел(?:а|и|ось)\s+бы\s+записат(?:ь|ься)\b|"
    r"\bможно\s+(?:сегодня\s+|на\s+сегодня\s+|на\s+\d{1,2}[./]\d{1,2}\s+)?записат(?:ься|ь)\b"
    r")",
    re.IGNORECASE,
)

MK_NAME_AFTER_TRIGGER_RE = re.compile(
    r"(?:"
    r"запиш(?:іть|и|ите)|"
    r"записат(?:и|ь)|"
    r"прошу\s+записат(?:и|ь)|"
    r"хочу\s+записат(?:и|ь)|"
    r"хотим\s+записат(?:ь)|"
    r"хочу\s+записать"
    r")"
    r"(?:\s+(?:будь\s+ласка|пожалуйста))?"
    r"\s*[:,-]?\s*(.*)$",
    re.IGNORECASE,
)


PHONE_RE = re.compile(
    r"(?<!\d)(?:\+?38[\s()\-]*0|0)[\s()\-]*\d{2}"
    r"(?:[\s()\-]*\d){7}(?!\d)"
)


def normalize_phone(phone):
    digits = re.sub(r"\D", "", phone or "")
    if digits.startswith("380") and len(digits) == 12:
        return "+" + digits
    if digits.startswith("0") and len(digits) == 10:
        return "+38" + digits
    return "+" + digits if digits else ""


def extract_phone(text):
    match = PHONE_RE.search(text or "")
    if not match:
        return None, text or ""
    phone = normalize_phone(match.group(0))
    text_without_phone = ((text or "")[:match.start()] + " " + (text or "")[match.end():]).strip()
    text_without_phone = re.sub(r"\s{2,}", " ", text_without_phone)
    text_without_phone = text_without_phone.strip(" ,;:-")
    return phone, text_without_phone


def get_saved_mk_phone(user_id):
    with sqlite3.connect(DATABASE) as conn:
        row = conn.execute(
            "SELECT phone_number FROM mk_contacts WHERE user_id=?",
            (user_id,),
        ).fetchone()
    return row[0] if row else None


def save_mk_phone(user_id, phone_number):
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            INSERT INTO mk_contacts(user_id, phone_number, updated_at)
            VALUES(?,?,?)
            ON CONFLICT(user_id) DO UPDATE SET
                phone_number=excluded.phone_number,
                updated_at=excluded.updated_at
            """,
            (
                user_id,
                phone_number,
                datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        conn.commit()


def clean_mk_name(raw_name):
    return raw_name.strip().strip('"\'«»„“”').strip()


BOOKING_SERVICE_WORDS_RE = re.compile(
    r"^(?:"
    r"будь\s+ласка|пожалуйста|"
    r"мене|меня|нас|"
    r"дитину|ребенка|ребёнка|"
    r"на\s+сьогодні|на\s+сегодня|сьогодні|сегодня|"
    r"на\s+завтра|завтра|"
    r"на\s+майстер[-\s]?клас|на\s+мастер[-\s]?класс|"
    r"на\s+мк|"
    r"на\s+\d{1,2}[./]\d{1,2}(?:[./]\d{2,4})?|"
    r"на\s+цей\s+майстер[-\s]?клас|на\s+этот\s+мастер[-\s]?класс"
    r")"
    r"(?:\s*[?.!,]*)$",
    re.IGNORECASE,
)


def extract_child_name_from_booking_text(text):
    """Витягує тільки ім'я/імена дітей із фрази запису."""
    value = str(text or "").strip()
    if not value:
        return ""

    _, value = extract_phone(value)

    # Привітання.
    value = re.sub(
        r"^\s*(?:доброго\s+(?:дня|ранку|вечора)|добрий\s+(?:день|вечір)|"
        r"доброе\s+утро|добрый\s+(?:день|вечер)|вітаю|здравствуйте|привет)"
        r"\s*[,!.:\-–—]*\s*",
        "",
        value,
        flags=re.IGNORECASE,
    )

    # Тригер запису.
    value = re.sub(
        r"^\s*(?:"
        r"хочу\s+записат(?:и|ись|ися|ь|ься)|"
        r"хотіл(?:а|и)?\s+б\s+записат(?:и|ись|ися)|"
        r"хотел(?:а|и|ось)?\s+бы\s+записат(?:ь|ься)|"
        r"можна\s+(?:мене\s+|нас\s+|дитину\s+|дітей\s+)?записат(?:и|ись|ися)|"
        r"можно\s+(?:меня\s+|нас\s+|ребенка\s+|ребёнка\s+|детей\s+)?записат(?:ь|ься)|"
        r"запиш(?:іть|и|ите|ем|емо)|записат(?:и|ь)|"
        r"прошу\s+записат(?:и|ь)"
        r")\s*[,!.:\-–—]*\s*",
        "",
        value,
        flags=re.IGNORECASE,
    )

    # Дата/час/МК.
    for pattern in (
        r"\bна\s+сьогодні\b", r"\bна\s+сегодня\b",
        r"\bсьогодні\b", r"\bсегодня\b",
        r"\bна\s+завтра\b", r"\bзавтра\b",
        r"\bна\s+неділю\b", r"\bна\s+воскресенье\b",
        r"\bна\s+майстер[-\s]?клас\b", r"\bна\s+мастер[-\s]?класс\b",
        r"\bна\s+мк\b",
        r"\bна\s+\d{1,2}[./]\d{1,2}(?:[./]\d{2,4})?\b",
    ):
        value = re.sub(pattern, " ", value, flags=re.IGNORECASE)

    # Ввічливість.
    value = re.sub(
        r"\b(?:будь\s*[- ]?\s*ласка|пожалуйста|прошу|дякую|спасибо)\b",
        " ",
        value,
        flags=re.IGNORECASE,
    )

    # Кількість дітей.
    value = re.sub(
        r"\b(?:\d{1,2}|одна|один|два|дві|две|двоє|двое|"
        r"три|троє|трое|чотири|четыре|четверо|"
        r"п['ʼ]?ять|пять|п['ʼ]?ятеро|пятеро)\s+"
        r"(?:дитину|дитини|дітей|діток|диток|ребенка|ребёнка|детей)\b",
        " ",
        value,
        flags=re.IGNORECASE,
    )

    # Залишкові службові слова.
    value = re.sub(
        r"\b(?:дитину|дітей|діток|ребенка|ребёнка|детей|нас|мене|меня)\b",
        " ",
        value,
        flags=re.IGNORECASE,
    )

    # Питання без імені.
    if re.search(
        r"\b(?:чи\s+записали|записали\s+дітей|записали\s+детей|"
        r"можна\s+записатися|можно\s+записаться)\b",
        value,
        flags=re.IGNORECASE,
    ):
        return ""

    value = re.sub(r"\s+", " ", value)
    value = value.strip(" ,;:.!?()[]{}'\"«»„“”")
    value = re.sub(r"\s*[,;]\s*", ", ", value).strip()

    if not value or BOOKING_SERVICE_WORDS_RE.fullmatch(value):
        return ""

    return clean_mk_name(value)


def get_mk_settings(group_id):
    with sqlite3.connect(DATABASE) as conn:
        row = conn.execute(
            """
            SELECT response_text, confirmation_text
            FROM mk_settings
            WHERE group_id=?
            """,
            (group_id,),
        ).fetchone()

    if row:
        return row
    return DEFAULT_MK_RESPONSE, DEFAULT_MK_CONFIRMATION


def format_mk_text(template, child_name, user_name, group_title):
    try:
        return template.format(
            name=child_name,
            user=user_name,
            group=group_title,
        )
    except (KeyError, ValueError):
        return template


async def start_mk_setting(update, setting_type):
    """Запускає приватний покроковий майстер налаштування тексту МК."""
    if not can_manage_mk(update.effective_user.id):
        return

    if update.effective_chat.type != "private":
        await update.message.reply_text(
            "🔒 Налаштування відповідей виконується тільки в особистому чаті з ботом."
        )
        return

    mk_config_state[update.effective_user.id] = {
        "step": "waiting_text",
        "setting_type": setting_type,
    }

    if setting_type == "response":
        title = "першої відповіді після фрази «Записати Ім’я»"
    else:
        title = "підтвердження після успішного запису"

    await update.message.reply_text(
        f"✏️ Надішліть текст для {title}.\n\n"
        "Можна використовувати змінні:\n"
        "{name} — ім’я дитини\n"
        "{user} — ім’я користувача Telegram\n"
        "{group} — назва групи\n\n"
        "Для скасування: /cancelmk"
    )


async def set_mk_response(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await start_mk_setting(update, "response")


async def set_mk_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await start_mk_setting(update, "confirmation")


async def cancel_mk_setting(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return

    mk_config_state.pop(update.effective_user.id, None)
    await update.message.reply_text("❌ Налаштування МК скасовано.")


async def receive_mk_setting_text(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    """Отримує приватний текст і пропонує обрати групу МК."""
    user = update.effective_user
    message = update.effective_message

    if not user or not message or not message.text:
        return
    if not can_manage_mk(user.id) or update.effective_chat.type != "private":
        return

    # Під час керування ролями приватний текст не є налаштуванням МК.
    if user.id in access_state:
        return

    state = mk_config_state.get(user.id)
    if not state or state.get("step") != "waiting_text":
        return

    phrase = get_message_html(message).strip()
    if not phrase:
        await message.reply_text("Текст не може бути порожнім. Надішліть фразу ще раз.")
        return

    groups_list = get_accessible_groups(user.id)
    if not groups_list:
        mk_config_state.pop(user.id, None)
        await message.reply_text("❌ Немає збережених груп. Спочатку додайте бота до групи.")
        return

    state["step"] = "waiting_group"
    state["phrase"] = phrase

    setting_type = state["setting_type"]
    keyboard = []

    if is_global_mk_role(user.id):
        keyboard.append(
            [
                InlineKeyboardButton(
                    "🌐 Усі групи",
                    callback_data=f"mkcfg:{setting_type}:all",
                )
            ]
        )

    for group_id, title in groups_list:
        button_title = title if len(title) <= 48 else title[:45] + "..."
        keyboard.append(
            [
                InlineKeyboardButton(
                    button_title,
                    callback_data=f"mkcfg:{setting_type}:{group_id}",
                )
            ]
        )

    keyboard.append(
        [InlineKeyboardButton("❌ Скасувати", callback_data="mkcfg:cancel")]
    )

    await message.reply_text(
        "📂 Для якої групи використати цю фразу?",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


async def mk_setting_buttons(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    """Зберігає налаштовану фразу для вибраної групи."""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not can_manage_mk(user_id):
        await query.answer("Немає доступу.", show_alert=True)
        return

    if query.data == "mkcfg:cancel":
        mk_config_state.pop(user_id, None)
        await query.edit_message_text("❌ Налаштування МК скасовано.")
        return

    state = mk_config_state.get(user_id)
    if not state or state.get("step") != "waiting_group":
        await query.edit_message_text(
            "❌ Сеанс налаштування завершився. Запустіть команду ще раз."
        )
        return

    try:
        _, setting_type, group_id_text = query.data.split(":", 2)
    except (ValueError, AttributeError):
        await query.edit_message_text("❌ Некоректні дані групи.")
        return

    phrase = state["phrase"]

    if group_id_text == "all":
        if not is_global_mk_role(user_id):
            await query.edit_message_text("❌ Немає доступу до всіх груп.")
            return

        groups_list = get_groups()
        if not groups_list:
            await query.edit_message_text("❌ Немає збережених груп.")
            return

        with sqlite3.connect(DATABASE) as conn:
            for group_id, _title in groups_list:
                response_text, confirmation_text = get_mk_settings(group_id)

                if setting_type == "response":
                    response_text = phrase
                elif setting_type == "confirmation":
                    confirmation_text = phrase
                else:
                    await query.edit_message_text("❌ Невідомий тип налаштування.")
                    return

                conn.execute(
                    """
                    INSERT INTO mk_settings(group_id, response_text, confirmation_text)
                    VALUES(?,?,?)
                    ON CONFLICT(group_id) DO UPDATE SET
                        response_text=excluded.response_text,
                        confirmation_text=excluded.confirmation_text
                    """,
                    (group_id, response_text, confirmation_text),
                )

            conn.commit()

        mk_config_state.pop(user_id, None)
        await query.edit_message_text(
            "✅ Фразу збережено для всіх груп."
        )
        return

    try:
        group_id = int(group_id_text)
    except ValueError:
        await query.edit_message_text("❌ Некоректні дані групи.")
        return

    if not can_access_group(user_id, group_id):
        await query.edit_message_text("❌ Немає доступу до цієї групи.")
        return

    response_text, confirmation_text = get_mk_settings(group_id)

    if setting_type == "response":
        response_text = phrase
        setting_name = "відповідь на запит запису"
    elif setting_type == "confirmation":
        confirmation_text = phrase
        setting_name = "підтвердження запису"
    else:
        await query.edit_message_text("❌ Невідомий тип налаштування.")
        return

    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            INSERT INTO mk_settings(group_id, response_text, confirmation_text)
            VALUES(?,?,?)
            ON CONFLICT(group_id) DO UPDATE SET
                response_text=excluded.response_text,
                confirmation_text=excluded.confirmation_text
            """,
            (group_id, response_text, confirmation_text),
        )
        group_row = conn.execute(
            "SELECT title FROM groups WHERE group_id=?",
            (group_id,),
        ).fetchone()
        conn.commit()

    group_title = group_row[0] if group_row else str(group_id)
    mk_config_state.pop(user_id, None)

    await query.edit_message_text(
        f"✅ Збережено: {html.escape(setting_name, quote=False)}.\n\n"
        f"🏷 Група: {html.escape(group_title, quote=False)}\n\n"
        f"💬 Фраза:\n{phrase}",
        parse_mode="HTML",
    )


async def _complete_mk_registration(
    message, chat, user, child_name, phone_number, context
):
    register_group(chat)
    user_name = user.full_name or str(user.id)
    created_at = datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
    phone_number = normalize_phone(phone_number)

    save_mk_phone(user.id, phone_number)

    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            INSERT INTO mk_registrations(
                group_id, group_title, requester_id, requester_name,
                child_name, phone_number, created_at
            )
            VALUES(?,?,?,?,?,?,?)
            """,
            (
                chat.id, chat.title, user.id, user_name,
                child_name, phone_number, created_at,
            ),
        )
        conn.commit()

    # Якщо номер був написаний у групі, прибираємо повідомлення з номером.
    if message and message.text and PHONE_RE.search(message.text):
        try:
            await context.bot.delete_message(
                chat_id=message.chat_id,
                message_id=message.message_id,
            )
        except Exception:
            logging.exception("Не вдалося видалити повідомлення з номером у групі")

    _, confirmation_text = get_mk_settings(chat.id)
    confirmation_text = format_mk_html(
        confirmation_text, child_name, user_name, chat.title or ""
    )
    try:
        await context.bot.send_message(
            chat_id=chat.id,
            text=confirmation_text,
            parse_mode="HTML",
        )
    except Exception:
        logging.exception("Не вдалося надіслати підтвердження в групу %s", chat.id)



async def _create_mk_request(
    message, chat, user, child_name, context, phone_number=None
):
    # Використовуємо номер із повідомлення або раніше збережений за Telegram ID.
    phone_number = phone_number or get_saved_mk_phone(user.id)
    if phone_number:
        await _complete_mk_registration(
            message, chat, user, child_name, phone_number, context
        )
        return

    register_group(chat)
    token = secrets.token_urlsafe(18)
    user_name = user.full_name or str(user.id)

    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            INSERT INTO mk_pending(
                token, group_id, group_title, requester_id, requester_name,
                child_name, source_message_id, created_at, status
            )
            VALUES(?,?,?,?,?,?,?,?,?)
            """,
            (
                token, chat.id, chat.title, user.id, user_name, child_name,
                message.message_id,
                datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S"),
                "waiting_start",
            ),
        )
        conn.commit()

    bot_info = await context.bot.get_me()
    private_url = f"https://t.me/{bot_info.username}?start=mk_{token}"
    response_text, _ = get_mk_settings(chat.id)
    response_text = format_mk_html(
        response_text, child_name, user_name, chat.title or ""
    )

    keyboard = InlineKeyboardMarkup(
        [[InlineKeyboardButton("📱 Надіслати номер телефону", url=private_url)]]
    )
    await message.reply_text(
        response_text,
        reply_markup=keyboard,
        parse_mode="HTML",
    )


async def mk_registration_trigger(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.effective_message
    chat = update.effective_chat
    user = update.effective_user

    if not message or not message.text or not user:
        return
    if chat.type not in ("group", "supergroup"):
        return

    state_key = (chat.id, user.id)

    # Якщо бот уже попросив ім'я, наступне текстове повідомлення
    # цього користувача в цій групі вважаємо ім'ям/іменами дітей.
    if state_key in mk_name_wait:
        state = mk_name_wait.get(state_key) or {}
        phone_number, name_text = extract_phone(message.text)
        phone_number = phone_number or state.get("phone_number")
        child_name = clean_mk_name(name_text)

        if not child_name:
            temp_message = await message.reply_text(
                "Підкажіть, будь ласка, ім'я дитини 👶"
            )
            schedule_message_deletion(
                context,
                chat.id,
                temp_message.message_id,
                delay=120,
            )
            return

        mk_name_wait.pop(state_key, None)

        # Видаляємо тільки службове питання бота.
        # Повідомлення гостя залишаємо в групі.
        await safe_delete_message(
            context,
            chat.id,
            state.get("prompt_message_id"),
        )

        await _create_mk_request(
            message,
            chat,
            user,
            child_name,
            context,
            phone_number,
        )
        return

    # Реагуємо на будь-який зрозумілий нам намір записатися.
    if not MK_TRIGGER_RE.search(message.text):
        return

    phone_number, text_without_phone = extract_phone(message.text)
    child_name = extract_child_name_from_booking_text(text_without_phone)

    # Якщо людина просто попросила записати, але не назвала дитину —
    # обов'язково уточнюємо ім'я перед вибором майстер-класу.
    if not child_name:
        prompt_message = await message.reply_text(
            "Підкажіть, будь ласка, ім'я дитини 👶"
        )

        mk_name_wait[state_key] = {
            "prompt_message_id": prompt_message.message_id,
            "trigger_message_id": message.message_id,
            "phone_number": phone_number,
        }

        schedule_message_deletion(
            context,
            chat.id,
            prompt_message.message_id,
            delay=120,
        )

        # Повідомлення гостя НЕ видаляємо.
        return

    await _create_mk_request(
        message,
        chat,
        user,
        child_name,
        context,
        phone_number,
    )


async def begin_mk_private_registration(update, context, token):
    message = update.effective_message
    user = update.effective_user
    chat = update.effective_chat

    if chat.type != "private":
        return

    with sqlite3.connect(DATABASE) as conn:
        pending = conn.execute(
            """
            SELECT group_id, group_title, requester_id, requester_name, child_name, status
            FROM mk_pending
            WHERE token=?
            """,
            (token,),
        ).fetchone()

        if not pending:
            await message.reply_text("❌ Посилання недійсне або запис уже завершено.")
            return

        group_id, group_title, requester_id, requester_name, child_name, status = pending
        if requester_id != user.id:
            await message.reply_text("❌ Це посилання створене для іншого користувача.")
            return
        if status == "completed":
            await message.reply_text("✅ Цей запис уже завершено.")
            return

        saved_phone = get_saved_mk_phone(user.id)
        if saved_phone:
            created_at = datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                """
                INSERT INTO mk_registrations(
                    group_id, group_title, requester_id, requester_name,
                    child_name, phone_number, created_at
                )
                VALUES(?,?,?,?,?,?,?)
                """,
                (
                    group_id, group_title, user.id, requester_name,
                    child_name, saved_phone, created_at,
                ),
            )
            conn.execute(
                "UPDATE mk_pending SET status='completed' WHERE token=?",
                (token,),
            )
            conn.commit()
        else:
            created_at = None

        if saved_phone:
            # Завершення поза транзакцією: повідомлення в групу та адміністраторам.
            pass
        else:
            conn.execute(
                "UPDATE mk_pending SET status='waiting_contact' WHERE token=?",
                (token,),
            )
            conn.execute(
                """
                INSERT INTO mk_user_state(user_id, token)
                VALUES(?,?)
                ON CONFLICT(user_id) DO UPDATE SET token=excluded.token
                """,
                (user.id, token),
            )
            conn.commit()

    if saved_phone:
        _, confirmation_text = get_mk_settings(group_id)
        confirmation_text = format_mk_html(
            confirmation_text, child_name, requester_name, group_title or ""
        )
        try:
            await context.bot.send_message(
            chat_id=group_id,
            text=confirmation_text,
            parse_mode="HTML",
        )
        except Exception:
            logging.exception("Не вдалося надіслати підтвердження в групу %s", group_id)

        await message.reply_text(
            f"✅ Запис для «{child_name}» створено. Використано збережений номер."
        )
        return

    keyboard = ReplyKeyboardMarkup(
        [[KeyboardButton("📱 Надіслати свій номер", request_contact=True)]],
        resize_keyboard=True,
        one_time_keyboard=True,
        input_field_placeholder="Натисніть кнопку нижче",
    )
    await message.reply_text(
        f"Запис для «{child_name}» у групі «{group_title}».\n\n"
        "Telegram не дозволяє боту автоматично прочитати прихований номер. "
        "Натисніть кнопку та поділіться своїм номером.",
        reply_markup=keyboard,
    )


async def receive_mk_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.effective_message
    user = update.effective_user
    contact = message.contact if message else None

    if not message or not user or not contact:
        return
    if update.effective_chat.type != "private":
        return

    # Не приймаємо чужу контактну картку замість номера автора запиту.
    if contact.user_id is not None and contact.user_id != user.id:
        await message.reply_text("❌ Будь ласка, надішліть саме свій номер кнопкою нижче.")
        return

    with sqlite3.connect(DATABASE) as conn:
        state = conn.execute(
            "SELECT token FROM mk_user_state WHERE user_id=?",
            (user.id,),
        ).fetchone()

        if not state:
            await message.reply_text(
                "Немає активного запису. Поверніться до групи та напишіть «Записати Ім’я».",
                reply_markup=ReplyKeyboardRemove(),
            )
            return

        token = state[0]
        pending = conn.execute(
            """
            SELECT group_id, group_title, requester_name, child_name
            FROM mk_pending
            WHERE token=? AND requester_id=? AND status='waiting_contact'
            """,
            (token, user.id),
        ).fetchone()

        if not pending:
            conn.execute("DELETE FROM mk_user_state WHERE user_id=?", (user.id,))
            conn.commit()
            await message.reply_text(
                "❌ Активний запис не знайдено.",
                reply_markup=ReplyKeyboardRemove(),
            )
            return

        group_id, group_title, requester_name, child_name = pending
        created_at = datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
        phone_number = normalize_phone(contact.phone_number)

        conn.execute(
            """
            INSERT INTO mk_registrations(
                group_id,
                group_title,
                requester_id,
                requester_name,
                child_name,
                phone_number,
                created_at
            )
            VALUES(?,?,?,?,?,?,?)
            """,
            (
                group_id,
                group_title,
                user.id,
                requester_name,
                child_name,
                phone_number,
                created_at,
            ),
        )
        conn.execute("UPDATE mk_pending SET status='completed' WHERE token=?", (token,))
        conn.execute(
            """
            INSERT INTO mk_contacts(user_id, phone_number, updated_at)
            VALUES(?,?,?)
            ON CONFLICT(user_id) DO UPDATE SET
                phone_number=excluded.phone_number,
                updated_at=excluded.updated_at
            """,
            (user.id, phone_number, created_at),
        )
        conn.execute("DELETE FROM mk_user_state WHERE user_id=?", (user.id,))
        conn.commit()

    # Повідомлення користувача з контактом не видаляємо.

    await context.bot.send_message(
        chat_id=user.id,
        text=f"✅ Запис для «{child_name}» успішно створено.",
        reply_markup=ReplyKeyboardRemove(),
    )

    _, confirmation_text = get_mk_settings(group_id)
    confirmation_text = format_mk_html(
        confirmation_text,
        child_name,
        requester_name,
        group_title or "",
    )
    try:
        await context.bot.send_message(
            chat_id=group_id,
            text=confirmation_text,
            parse_mode="HTML",
        )
    except Exception:
        logging.exception("Не вдалося надіслати підтвердження в групу %s", group_id)



async def mk_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return

    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(
            """
            SELECT group_title, child_name, requester_name, phone_number, created_at
            FROM mk_registrations
            ORDER BY id DESC
            LIMIT 20
            """
        ).fetchall()

    if not rows:
        await update.message.reply_text("Записів на майстер-класи ще немає.")
        return

    parts = ["📝 Останні записи на МК:\n"]
    for group_title, child_name, requester_name, phone_number, created_at in rows:
        parts.append(
            f"\n🍣 {group_title}\n"
            f"Ім’я: {child_name}\n"
            f"Користувач: {requester_name}\n"
            f"Телефон: {phone_number}\n"
            f"Дата: {created_at}\n"
        )

    text = "".join(parts)
    for start in range(0, len(text), 4000):
        await update.message.reply_text(text[start:start + 4000])


# ============================================================
# РОЗСИЛКА З ПІДТВЕРДЖЕННЯМ
# ============================================================


async def broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    broadcast_wait[update.effective_user.id] = True
    await update.message.reply_text('📨 Надішліть повідомлення для розсилки.\n\nПідтримуються:\n• текст\n• Premium Emoji\n• фото\n• відео\n• GIF\n• документ\n\nПісля цього бот покаже кнопки підтвердження.')


async def receive_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id not in broadcast_wait:
        return
    del broadcast_wait[user_id]
    message = update.message
    broadcast_message[user_id] = (message.chat.id, message.message_id)
    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton('✅ Відправити', callback_data='broadcast_send'), InlineKeyboardButton('❌ Скасувати', callback_data='broadcast_cancel')]])
    await update.message.reply_text('👆 Повідомлення отримано.\n\nПідтвердити розсилку?', reply_markup=keyboard)


async def broadcast_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if query.data == 'broadcast_cancel':
        if user_id in broadcast_message:
            del broadcast_message[user_id]
        await query.edit_message_text('❌ Розсилку скасовано.')
        return
    if query.data != 'broadcast_send':
        return
    if user_id not in broadcast_message:
        await query.edit_message_text('❌ Повідомлення не знайдено.')
        return
    from_chat_id, message_id = broadcast_message[user_id]
    groups = get_groups()
    sent = 0
    failed = 0
    for group_id, title in groups:
        try:
            msg = await context.bot.copy_message(chat_id=group_id, from_chat_id=from_chat_id, message_id=message_id)
            with sqlite3.connect(DATABASE) as conn:
                conn.execute('\n                    INSERT INTO last_broadcast(\n                        group_id,\n                        message_id\n                    )\n                    VALUES (?, ?)\n                    ', (group_id, msg.message_id))
                conn.commit()
            sent += 1
        except Exception as e:
            print(f'{title}: {e}')
            failed += 1
    del broadcast_message[user_id]
    await query.edit_message_text(f'\n✅ Розсилку завершено\n\n📨 Надіслано: {sent}\n❌ Помилок: {failed}\n📂 Всього груп: {len(groups)}\n')


# ============================================================
# ВИДАЛЕННЯ ПОВІДОМЛЕНЬ
# ============================================================


async def delete_by_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    if len(context.args) != 2:
        await update.message.reply_text('❌ Формат:\n/delete_id CHAT_ID MESSAGE_ID\n\nПример:\n/delete_id -1001234567890 456')
        return
    try:
        chat_id = int(context.args[0])
        message_id = int(context.args[1])
        await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
        await update.message.reply_text('🗑 Сообщение удалено.')
    except Exception as e:
        print('Ошибка удаления:', e)
        await update.message.reply_text(f'❌ Ошибка:\n{e}')


async def deletebroadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute('\n            SELECT group_id, message_id\n            FROM last_broadcast\n            ').fetchall()
    if not rows:
        await update.message.reply_text('Немає повідомлень для видалення.')
        return
    deleted = 0
    failed = 0
    for group_id, message_id in rows:
        try:
            await context.bot.delete_message(chat_id=group_id, message_id=message_id)
            deleted += 1
        except Exception as e:
            print(e)
            failed += 1
    with sqlite3.connect(DATABASE) as conn:
        conn.execute('DELETE FROM last_broadcast')
        conn.commit()
    await update.message.reply_text(f'\n🗑 Розсилку видалено\n\n✅ Видалено: {deleted}\n❌ Помилок: {failed}\n')


async def delete_forwarded_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        return
    if user_id not in delete_message_cache:
        await update.message.reply_text('❌ Сначала перешли сообщение.')
        return
    chat_id, message_id = delete_message_cache[user_id]
    try:
        await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
        await update.message.reply_text('🗑 Сообщение удалено.')
        del delete_message_cache[user_id]
    except Exception as e:
        print(e)
        await update.message.reply_text(f'❌ Ошибка удаления:\n{e}')


# ============================================================
# ОБРОБНИК ПОМИЛОК
# ============================================================


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logging.error('Помилка:', exc_info=context.error)


# ============================================================
# ДОПОМІЖНЕ ЗБЕРЕЖЕННЯ ГРУПИ
# ============================================================

async def save_group(update: Update):
    chat = update.effective_chat
    if chat.type not in ("group", "supergroup"):
        return
    register_group(chat)



# ============================================================
# ГРАФІК МАЙСТЕР-КЛАСІВ ТА ПОСТІЙНИЙ EXCEL ІЗ ЗАПИСАМИ
# ============================================================

MK_REGISTRATIONS_FILE = Path("Записи_на_майстер-класи.xlsx")
mk_schedule_state = {}
_mk_excel_lock = threading.RLock()

# Тимчасові повідомлення в групі автоматично видаляються.
TEMP_GROUP_MESSAGE_SECONDS = 90


async def delete_message_later(context, chat_id, message_id, delay=TEMP_GROUP_MESSAGE_SECONDS):
    """Видаляє службове повідомлення із групи через заданий час."""
    await asyncio.sleep(delay)
    try:
        await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
    except Exception:
        # Повідомлення вже могло бути видалене або бот не має права видалення.
        pass


def get_message_html(message):
    """Повертає Telegram HTML зі збереженням Premium/Custom Emoji."""
    if not message:
        return ""
    try:
        value = getattr(message, "text_html", None)
        if value:
            return value
    except Exception:
        pass
    try:
        value = getattr(message, "caption_html", None)
        if value:
            return value
    except Exception:
        pass
    return message.text or message.caption or ""


def format_mk_html(template, child_name, user_name, group_title):
    """Підставляє персональні дані в HTML-шаблон без втрати tg-emoji."""
    values = {
        "name": html.escape(str(child_name or ""), quote=False),
        "user": html.escape(str(user_name or ""), quote=False),
        "group": html.escape(str(group_title or ""), quote=False),
    }
    try:
        return template.format(**values)
    except (KeyError, ValueError, IndexError):
        return template


def is_global_mk_role(user_id):
    role = get_user_role(user_id)
    return role in {ROLE_DEVELOPER, ROLE_MANAGER}

def schedule_message_deletion(context, chat_id, message_id, delay=TEMP_GROUP_MESSAGE_SECONDS):
    context.application.create_task(
        delete_message_later(context, chat_id, message_id, delay),
        name=f"delete-message-{chat_id}-{message_id}",
    )


async def safe_delete_message(context, chat_id, message_id):
    if not message_id:
        return
    try:
        await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
    except Exception:
        pass


def _ensure_column(conn, table_name, column_name, definition):
    columns = {
        row[1] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    }
    if column_name not in columns:
        conn.execute(
            f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}"
        )


def init_mk_schedule_schema():
    """Додає таблиці та колонки МК без видалення старих даних."""
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS mk_schedule(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                group_id INTEGER NOT NULL,
                event_date TEXT NOT NULL,
                title TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(group_id, event_date, title)
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS mk_booking_drafts(
                token TEXT PRIMARY KEY,
                group_id INTEGER NOT NULL,
                group_title TEXT,
                requester_id INTEGER NOT NULL,
                requester_name TEXT,
                child_name TEXT NOT NULL,
                phone_number TEXT,
                source_message_id INTEGER,
                phone_was_in_group INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
            """
        )

        # ВАЖЛИВО: спочатку додаємо відсутні колонки,
        # і лише після цього будь-які функції можуть робити SELECT по них.
        _ensure_column(conn, "mk_pending", "event_date", "TEXT")
        _ensure_column(conn, "mk_pending", "event_title", "TEXT")
        _ensure_column(conn, "mk_pending", "prompt_message_id", "INTEGER")

        _ensure_column(conn, "mk_registrations", "event_date", "TEXT")
        _ensure_column(conn, "mk_registrations", "event_title", "TEXT")
        _ensure_column(conn, "mk_registrations", "attendance_status", "TEXT NOT NULL DEFAULT 'yes'")
        _ensure_column(conn, "mk_registrations", "google_source_key", "TEXT")

        # Стан вихідної синхронізації бот -> Google.
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS mk_google_sync(
                registration_id INTEGER PRIMARY KEY,
                synced INTEGER NOT NULL DEFAULT 0,
                attempts INTEGER NOT NULL DEFAULT 0,
                last_error TEXT,
                synced_at TEXT
            )
            """
        )

        # Один і той самий рядок Google не імпортуємо повторно.
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS
                idx_mk_registrations_google_source_key
            ON mk_registrations(google_source_key)
            WHERE google_source_key IS NOT NULL
            """
        )

        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=10000")
        conn.commit()

    # Excel оновлюємо лише після успішного завершення міграції схеми.
    try:
        refresh_mk_registrations_excel()
    except Exception:
        logging.exception("Не вдалося оновити Excel після міграції схеми МК")

def get_active_mk_schedule(group_id):
    today = datetime.now(TIMEZONE).date().isoformat()
    with sqlite3.connect(DATABASE) as conn:
        return conn.execute(
            """
            SELECT id, event_date, title
            FROM mk_schedule
            WHERE group_id=? AND event_date>=?
            ORDER BY event_date, id
            LIMIT 5
            """,
            (group_id, today),
        ).fetchall()


def parse_mk_schedule(text: str):
    import re
    from datetime import datetime

    date_pattern = re.compile(r"(?m)^\s*(\d{2}\.\d{2}\.\d{4})(?:\s+([^\n]+))?\s*$")
    matches=list(date_pattern.finditer(text))

    if not 1<=len(matches)<=5:
        raise ValueError("Потрібно вказати від 1 до 5 майстер-класів.")

    entries=[]
    months=set()

    for i,match in enumerate(matches):
        date_text=match.group(1)
        first=(match.group(2) or "").strip()

        block=text[match.end():matches[i+1].start() if i+1<len(matches) else len(text)]
        lines=[x.strip() for x in block.splitlines() if x.strip()]

        dishes=[]
        if first:
            dishes.append(first)
        dishes.extend(lines)

        if len(dishes)!=2:
            raise ValueError(f"Для дати {date_text} потрібно вказати дві строки.")

        d=datetime.strptime(date_text,"%d.%m.%Y").date()
        if d.weekday()!=6:
            raise ValueError(f"{date_text} — не неділя.")

        months.add((d.year,d.month))
        entries.append({
            "date":d.strftime("%Y-%m-%d"),
            "date_display":d.strftime("%d.%m.%Y"),
            "title":dishes[0]+"\n"+dishes[1],
            "first_dish":dishes[0],
            "second_dish":dishes[1],
        })

    if len(months)!=1:
        raise ValueError("Усі дати повинні належати одному місяцю.")

    entries.sort(key=lambda x:x["date"])
    return entries

def make_unique_sheet_title(group_title, used_titles):
    """Створює коректну унікальну назву листа Excel."""
    invalid_chars = r'[]:*?/\\'
    title = str(group_title or "Без назви")

    for char in invalid_chars:
        title = title.replace(char, " ")

    title = re.sub(r"\s+", " ", title).strip() or "Без назви"
    title = title[:31]

    candidate = title
    number = 2

    while candidate.lower() in used_titles:
        suffix = f" ({number})"
        candidate = f"{title[:31 - len(suffix)]}{suffix}"
        number += 1

    used_titles.add(candidate.lower())
    return candidate


def format_mk_excel_date(value, input_format, output_format, default=""):
    """Безпечно форматує дату для Excel."""
    if not value:
        return default

    try:
        return datetime.strptime(value, input_format).strftime(output_format)
    except (TypeError, ValueError):
        return value

def _parse_excel_datetime(value, value_format):
    if not value:
        return None
    try:
        return datetime.strptime(value, value_format)
    except (TypeError, ValueError):
        return None


CHILD_COUNT_WORDS = {
    "одна": 1, "один": 1,
    "два": 2, "дві": 2, "две": 2, "двоє": 2, "двое": 2,
    "три": 3, "троє": 3, "трое": 3,
    "чотири": 4, "четыре": 4, "четверо": 4,
    "п'ять": 5, "пʼять": 5, "пять": 5,
    "п'ятеро": 5, "пʼятеро": 5, "пятеро": 5,
}

CHILD_WORD_PATTERN = (
    r"(?:ребёнка|ребенка|ребёнок|ребенок|"
    r"дитини|дитина|дітей|детей|дети|дит|реб)"
)


def extract_declared_children_count(text):
    value = str(text or "").strip()
    if not value:
        return None

    digit_match = re.search(
        rf"(?<!\w)(\d{{1,2}})\s*{CHILD_WORD_PATTERN}(?!\w)",
        value,
        re.IGNORECASE,
    )
    if digit_match:
        return max(1, int(digit_match.group(1)))

    lowered = value.casefold()
    for word, count in CHILD_COUNT_WORDS.items():
        if re.search(
            rf"(?<!\w){re.escape(word)}(?:\s+{CHILD_WORD_PATTERN})?(?!\w)",
            lowered,
            re.IGNORECASE,
        ):
            return count

    return None


def clean_children_names_text(child_names):
    value = str(child_names or "").strip()
    if not value:
        return ""

    number_word = (
        r"(?:\d{1,2}|одна|один|два|дві|две|двоє|двое|"
        r"три|троє|трое|чотири|четыре|четверо|"
        r"п['ʼ]?ять|пять|п['ʼ]?ятеро|пятеро)"
    )

    value = re.sub(
        rf"^\s*{number_word}\s*{CHILD_WORD_PATTERN}\s*[,;:\-–—]*\s*",
        "",
        value,
        flags=re.IGNORECASE,
    ).strip()

    return value


def split_children_names(child_names):
    value = clean_children_names_text(child_names)
    if not value:
        return []

    parts = re.split(
        r"\s*(?:[,;\n]+|\s+(?:і|й|та|и|and)\s+|\s*[+&]\s*)\s*",
        value,
        flags=re.IGNORECASE,
    )
    return [part.strip() for part in parts if part and part.strip()]


def normalized_children_names(child_names):
    """У Google childrenNames відправляємо тільки ім'я/імена."""
    value = str(child_names or "").strip()
    if not value:
        return ""

    cleaned = extract_child_name_from_booking_text(value)

    # Якщо сюди вже прийшло просте ім'я без тригерної фрази.
    if not cleaned:
        cleaned = clean_children_names_text(value)

    cleaned = re.sub(
        r"^\s*(?:будь\s*[- ]?\s*ласка|пожалуйста)\s*[,;:\-–—]*\s*",
        "",
        cleaned,
        flags=re.IGNORECASE,
    ).strip()

    return cleaned


def _children_count(child_names):
    declared = extract_declared_children_count(child_names)
    if declared is not None:
        return declared

    parts = split_children_names(child_names)
    return max(1, len(parts))


def _read_saved_excel_settings(filepath):
    """Зберігає ручні значення «Працює» та «Ліміт» між оновленнями файлу."""
    settings = {}
    if not filepath.exists():
        return settings
    try:
        old_wb = load_workbook(filepath, read_only=True, data_only=True)
        if "Налаштування" not in old_wb.sheetnames:
            old_wb.close()
            return settings
        ws = old_wb["Налаштування"]
        for row in ws.iter_rows(min_row=3, min_col=1, max_col=3, values_only=True):
            restaurant, active, limit = row
            if restaurant:
                settings[str(restaurant)] = {
                    "active": 1 if active in (None, "", True, 1, "1") else 0,
                    "limit": int(limit) if isinstance(limit, (int, float)) else 20,
                }
        old_wb.close()
    except Exception:
        logging.exception("Не вдалося прочитати попередні налаштування Excel")
    return settings


def _style_registration_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="B7B7B7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = (23, 38, 13, 24, 18, 12, 40)
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.column_dimensions["E"].number_format = "@"
    for cell in ws["A"][1:]:
        cell.number_format = "dd.mm.yyyy hh:mm"
    for cell in ws["C"][1:]:
        cell.number_format = "dd.mm.yyyy"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[4].number_format = "@"


def _build_mk_workbook(rows, groups, schedules, filepath):
    saved_settings = _read_saved_excel_settings(filepath)
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    registrations = []
    grouped = {}
    for row in rows:
        group_id, group_title, created_at, event_date, event_title, child_name, parent_name, phone = row
        title = group_title or str(group_id)
        record = {
            "group_id": group_id,
            "restaurant": title,
            "created": _parse_excel_datetime(created_at, "%Y-%m-%d %H:%M:%S"),
            "event_date": _parse_excel_datetime(event_date, "%Y-%m-%d"),
            "event_title": event_title or "",
            "parent": parent_name or "",
            "phone": str(phone or ""),
            "count": _children_count(child_name),
            "children": child_name or "",
        }
        registrations.append(record)
        grouped.setdefault((group_id, title), []).append(record)

    known_groups = {group_id: (title or str(group_id)) for group_id, title in groups}
    for group_id, title in grouped:
        known_groups.setdefault(group_id, title)
    dates = sorted({entry[1] for entry in schedules if entry[1]} | {
        r["event_date"].strftime("%Y-%m-%d") for r in registrations if r["event_date"]
    })

    dark_fill = PatternFill("solid", fgColor="274E13")
    light_fill = PatternFill("solid", fgColor="D9EAD3")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    white_bold = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B7B7B7")

    settings_ws = wb.create_sheet("Налаштування")
    settings_ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=max(4, 3 + len(dates)))
    settings_ws.append([])
    settings_ws["A1"], settings_ws["B1"], settings_ws["C1"], settings_ws["D1"] = "Ресторан", "Працює", "Ліміт", "Запис"
    settings_ws.merge_cells("A1:A2")
    settings_ws.merge_cells("B1:B2")
    settings_ws.merge_cells("C1:C2")
    for col, date_value in enumerate(dates, 4):
        settings_ws.cell(2, col, _parse_excel_datetime(date_value, "%Y-%m-%d"))
        settings_ws.cell(2, col).number_format = "dd.mm"
    for row_num, (group_id, title) in enumerate(sorted(known_groups.items(), key=lambda x: x[1].lower()), 3):
        config = saved_settings.get(title, {"active": 1, "limit": 20})
        settings_ws.cell(row_num, 1, title)
        settings_ws.cell(row_num, 2, config["active"])
        settings_ws.cell(row_num, 3, config["limit"])
        for col, date_value in enumerate(dates, 4):
            count = sum(r["count"] for r in registrations if r["group_id"] == group_id and r["event_date"] and r["event_date"].strftime("%Y-%m-%d") == date_value)
            settings_ws.cell(row_num, col, count)
    settings_ws.column_dimensions["A"].width = 43
    settings_ws.column_dimensions["B"].width = 10
    settings_ws.column_dimensions["C"].width = 10
    for col in range(4, 4 + len(dates)):
        settings_ws.column_dimensions[get_column_letter(col)].width = 9
    for row in settings_ws.iter_rows(min_row=1, max_row=2, max_col=max(4, 3 + len(dates))):
        for cell in row:
            cell.fill, cell.font = dark_fill, white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    settings_ws.freeze_panes = "D3"

    limit_ws = wb.create_sheet("Налаштування з лімітами")
    max_col = max(2, 1 + len(dates))
    limit_ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=max_col)
    limit_ws["A1"], limit_ws["B1"] = "Ресторан", "Записано дітей"
    for col, date_value in enumerate(dates, 2):
        limit_ws.cell(2, col, _parse_excel_datetime(date_value, "%Y-%m-%d"))
        limit_ws.cell(2, col).number_format = "dd.mm"
    group_items = sorted(known_groups.items(), key=lambda x: x[1].lower())
    for row_num, (group_id, title) in enumerate(group_items, 3):
        limit_ws.cell(row_num, 1, title)
        for col, date_value in enumerate(dates, 2):
            count = sum(r["count"] for r in registrations if r["group_id"] == group_id and r["event_date"] and r["event_date"].strftime("%Y-%m-%d") == date_value)
            limit_ws.cell(row_num, col, count)
    second_header = 4 + len(group_items)
    limit_ws.merge_cells(start_row=second_header, start_column=2, end_row=second_header, end_column=max_col)
    limit_ws.cell(second_header, 1, "Ресторан")
    limit_ws.cell(second_header, 2, "Залишилось місць")
    for col, date_value in enumerate(dates, 2):
        limit_ws.cell(second_header + 1, col, _parse_excel_datetime(date_value, "%Y-%m-%d"))
        limit_ws.cell(second_header + 1, col).number_format = "dd.mm"
    for offset, (group_id, title) in enumerate(group_items, second_header + 2):
        limit_ws.cell(offset, 1, title)
        limit = saved_settings.get(title, {"limit": 20})["limit"]
        for col, date_value in enumerate(dates, 2):
            count = sum(r["count"] for r in registrations if r["group_id"] == group_id and r["event_date"] and r["event_date"].strftime("%Y-%m-%d") == date_value)
            cell = limit_ws.cell(offset, col, max(0, limit - count))
            if count >= limit:
                cell.fill = red_fill
    limit_ws.column_dimensions["A"].width = 43
    for col in range(2, max_col + 1):
        limit_ws.column_dimensions[get_column_letter(col)].width = 10
    for header_row in (1, 2, second_header, second_header + 1):
        for cell in limit_ws[header_row][:max_col]:
            cell.fill, cell.font = dark_fill, white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    limit_ws.freeze_panes = "B3"

    headers = ["Дата/время регистрации", "Ресторан", "Дата", "Имя", "Телефон", "К-во детей", "Имена"]
    all_ws = wb.create_sheet("Всі реєстрації")
    all_ws.append(headers)
    for record in registrations:
        all_ws.append([record["created"], record["restaurant"], record["event_date"], record["parent"], record["phone"], record["count"], record["children"]])
    _style_registration_sheet(all_ws)

    used_titles = {name.lower() for name in wb.sheetnames}
    for group_id, title in group_items:
        ws = wb.create_sheet(make_unique_sheet_title(title, used_titles))
        ws.append(headers)
        for record in grouped.get((group_id, title), []):
            ws.append([record["created"], record["restaurant"], record["event_date"], record["parent"], record["phone"], record["count"], record["children"]])
        _style_registration_sheet(ws)

    total_ws = wb.create_sheet("Итого записей")
    total_ws.append(["Ресторан"] + [_parse_excel_datetime(d, "%Y-%m-%d") for d in dates] + ["Ліміт", "Працює"])
    for col in range(2, 2 + len(dates)):
        total_ws.cell(1, col).number_format = "dd.mm"
    for group_id, title in group_items:
        config = saved_settings.get(title, {"active": 1, "limit": 20})
        values = [title]
        for date_value in dates:
            values.append(sum(r["count"] for r in registrations if r["group_id"] == group_id and r["event_date"] and r["event_date"].strftime("%Y-%m-%d") == date_value))
        values.extend([config["limit"], config["active"]])
        total_ws.append(values)
    total_ws.column_dimensions["A"].width = 43
    for cell in total_ws[1]:
        cell.fill, cell.font = dark_fill, white_bold
        cell.alignment = Alignment(horizontal="center")
    total_ws.freeze_panes = "B2"

    lib_ws = wb.create_sheet("lib")
    for row_num, (_, title) in enumerate(group_items, 1):
        lib_ws.cell(row_num, 1, title)
    desc_ws = wb.create_sheet("desc")
    desc_ws.append(["Ресторан", "Дата", "Майстер-клас"])
    schedule_titles = {(g, d): title for g, d, title in schedules}
    for (group_id, date_value), title in sorted(schedule_titles.items(), key=lambda item: (item[0][1], item[0][0])):
        desc_ws.append([known_groups.get(group_id, str(group_id)), _parse_excel_datetime(date_value, "%Y-%m-%d"), title])
    desc_ws.column_dimensions["A"].width = 43
    desc_ws.column_dimensions["B"].width = 13
    desc_ws.column_dimensions["C"].width = 55
    desc_ws["B1"].number_format = "dd.mm.yyyy"
    for cell in desc_ws[1]:
        cell.fill, cell.font = light_fill, Font(bold=True)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and not cell.border.bottom.style:
                    cell.border = Border(bottom=thin)

    temp_path = filepath.with_name(f".{filepath.stem}.tmp{filepath.suffix}")
    wb.save(temp_path)
    temp_path.replace(filepath)
    return filepath


def build_mk_registrations_excel():
    """Атомарно перебудовує постійну книгу в структурі наданого зразка."""
    with _mk_excel_lock:
        with sqlite3.connect(DATABASE) as conn:
            rows = conn.execute(
                """
                SELECT group_id, group_title, created_at, event_date, event_title,
                       child_name, requester_name, phone_number
                FROM mk_registrations
                ORDER BY COALESCE(created_at, '' ) DESC, id DESC
                """
            ).fetchall()
            groups = conn.execute(
                "SELECT group_id, COALESCE(title, CAST(group_id AS TEXT)) FROM groups ORDER BY title"
            ).fetchall()
            schedules = conn.execute(
                "SELECT group_id, event_date, title FROM mk_schedule ORDER BY event_date, id"
            ).fetchall()
        return _build_mk_workbook(rows, groups, schedules, MK_REGISTRATIONS_FILE)


def refresh_mk_registrations_excel():
    """Оновлює звіт, але не ламає реєстрацію або запуск бота, якщо файл відкритий."""
    try:
        return build_mk_registrations_excel()
    except Exception:
        logging.exception("Не вдалося оновити постійний Excel із записами МК")
        return None


async def set_mk_schedule(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):

    if not can_manage_mk(update.effective_user.id):
        return

    if update.effective_chat.type != "private":

        await update.message.reply_text(
            "🔒 Графік налаштовується тільки "
            "в особистому чаті з ботом."
        )

        return

    mk_schedule_state[update.effective_user.id] = {
        "step": "waiting_schedule"
    }

    await update.message.reply_text(
        "📅 Надішліть графік на місяць одним повідомленням.\n\n"

        "Можна використовувати два формати.\n\n"

        "Варіант 1:\n"
        "09.08.2026\n"
        "БЕЛЬГІЙСЬКА ВАФЛЯ З МОРОЗИВОМ\n"
        "ДИТЯЧИЙ БАМБЛ\n\n"

        "Варіант 2:\n"
        "16.08.2026 БУРГЕР У ЛАВАШІ З КАРТОПЛЕЮ ФРІ\n"
        "ЛИМОНАД «СМАРАГДОВЕ НЕБО»\n\n"

        "Між майстер-класами бажано залишати "
        "порожній рядок.\n\n"

        "Дозволено від 1 до 5 недільних дат "
        "одного місяця.\n\n"

        "Для скасування: /cancelmk"
    )


async def receive_mk_schedule_text(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):

    user = update.effective_user
    message = update.effective_message

    if not user or not message or not message.text:
        return

    if (
        not is_admin(user.id)
        or update.effective_chat.type != "private"
    ):
        return

    # Під час керування ролями приватний текст не є графіком МК.
    if user.id in access_state:
        return

    state = mk_schedule_state.get(user.id)

    if (
        not state
        or state.get("step") != "waiting_schedule"
    ):
        return

    try:
        entries = parse_mk_schedule(message.text)

    except ValueError as exc:

        await message.reply_text(
            f"❌ {exc}\n\n"
            "Надішліть графік ще раз."
        )

        return

    groups_list = get_groups()

    if not groups_list:

        mk_schedule_state.pop(user.id, None)

        await message.reply_text(
            "❌ Немає збережених груп."
        )

        return

    state["step"] = "waiting_group"
    state["entries"] = entries

    keyboard = []

    # Розробник і Керуючий можуть застосувати графік одразу до всіх груп.
    if is_global_mk_role(user.id):
        keyboard.append(
            [
                InlineKeyboardButton(
                    "🌐 Усі групи",
                    callback_data="mkschedgrp:all",
                )
            ]
        )

    for group_id, title in groups_list:

        button_title = (
            title
            if len(title) <= 48
            else title[:45] + "..."
        )

        keyboard.append(
            [
                InlineKeyboardButton(
                    button_title,
                    callback_data=f"mkschedgrp:{group_id}",
                )
            ]
        )

    keyboard.append(
        [
            InlineKeyboardButton(
                "❌ Скасувати",
                callback_data="mkschedgrp:cancel",
            )
        ]
    )

    schedule_preview = "✅ Графік розпізнано:\n\n"

    for entry in entries:

        schedule_preview += (
            f"📅 {entry['date_display']}\n"
            f"🍣 {entry['first_dish']}\n"
            f"🥤 {entry['second_dish']}\n\n"
        )

    schedule_preview += (
        "📂 Для якої групи зберегти цей графік?"
    )

    await message.reply_text(
        schedule_preview,
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


async def mk_schedule_group_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if not can_manage_mk(user_id):
        await query.answer("Немає доступу.", show_alert=True)
        return

    if query.data == "mkschedgrp:cancel":
        mk_schedule_state.pop(user_id, None)
        await query.edit_message_text("❌ Налаштування графіка скасовано.")
        return

    if query.data == "mkschedgrp:all":
        if not is_global_mk_role(user_id):
            await query.edit_message_text("❌ Немає доступу до всіх груп.")
            return

        state = mk_schedule_state.get(user_id)
        if not state or state.get("step") != "waiting_group":
            await query.edit_message_text(
                "❌ Сеанс налаштування завершився. Запустіть команду ще раз."
            )
            return

        entries = state["entries"]
        if not entries:
            mk_schedule_state.pop(user_id, None)
            await query.edit_message_text("❌ Графік порожній.")
            return

        month_prefix = entries[0]["date"][:7]
        created_at = datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
        groups_list = get_groups()

        with sqlite3.connect(DATABASE) as conn:
            for group_id, _title in groups_list:
                conn.execute(
                    "DELETE FROM mk_schedule WHERE group_id=? AND substr(event_date,1,7)=?",
                    (group_id, month_prefix),
                )
                conn.executemany(
                    """
                    INSERT INTO mk_schedule(group_id, event_date, title, created_at)
                    VALUES(?,?,?,?)
                    """,
                    [
                        (
                            group_id,
                            entry["date"],
                            entry["title"],
                            created_at,
                        )
                        for entry in entries
                    ],
                )
            conn.commit()

        mk_schedule_state.pop(user_id, None)
        await query.edit_message_text(
            f"✅ Графік збережено для всіх груп.\n\n"
            f"Груп: {len(groups_list)}"
        )
        return

    state = mk_schedule_state.get(user_id)
    if not state or state.get("step") != "waiting_group":
        await query.edit_message_text("❌ Сеанс завершився. Запустіть /setmkschedule ще раз.")
        return

    try:
        group_id = int(query.data.split(":", 1)[1])
    except (ValueError, IndexError):
        await query.edit_message_text("❌ Некоректна група.")
        return

    if not can_access_group(user_id, group_id):
        await query.edit_message_text("❌ Немає доступу до цієї групи.")
        return

    entries = state["entries"]

    if not entries:
        mk_schedule_state.pop(user_id, None)
        await query.edit_message_text("❌ Графік порожній. Запустіть /setmkschedule ще раз.")
        return

    month_prefix = entries[0]["date"][:7]
    created_at = datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")

    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            "DELETE FROM mk_schedule WHERE group_id=? AND substr(event_date,1,7)=?",
            (group_id, month_prefix),
        )
        conn.executemany(
            """
            INSERT INTO mk_schedule(group_id, event_date, title, created_at)
            VALUES(?,?,?,?)
            """,
            [
                (
                    group_id,
                    entry["date"],
                    entry["title"],
                    created_at,
                )
                for entry in entries
            ],
        )
        group_row = conn.execute(
            "SELECT title FROM groups WHERE group_id=?",
            (group_id,),
        ).fetchone()
        conn.commit()

    mk_schedule_state.pop(user_id, None)
    group_title = group_row[0] if group_row else str(group_id)

    lines = [
        (
            f"• {entry['date_display']}\n"
            f"  🍣 {entry['first_dish']}\n"
            f"  🥤 {entry['second_dish']}"
        )
        for entry in entries
    ]
    await query.edit_message_text(
        f"✅ Графік збережено.\n\n🏷 {group_title}\n\n" + "\n".join(lines)
    )



def resolve_google_mk_restaurant(group_title):
    """Повертає назву ресторану так, як вона має виглядати в Google-таблиці."""
    title = (group_title or "").strip()
    normalized = title.casefold().replace("’", "'").replace("ʼ", "'")

    for fragment, restaurant in GOOGLE_MK_RESTAURANT_ALIASES.items():
        if fragment.casefold() in normalized:
            return restaurant

    # Fallback: endpoint все одно отримає непорожній rest і не втратить запис.
    # Для нових груп бажано додати окремий alias вище.
    return title


def infer_children_count(child_name):
    value = (child_name or "").strip()
    if not value:
        return 1

    declared = extract_declared_children_count(value)
    if declared is not None:
        return declared

    parts = split_children_names(value)
    return max(1, len(parts))


def validate_google_mk_payload(payload):
    required = ("name", "phone", "date", "rest", "children")
    missing = [key for key in required if not str(payload.get(key, "")).strip()]
    if missing:
        raise ValueError(
            "Не заповнені обов'язкові поля Google MK: " + ", ".join(missing)
        )


def build_google_mk_payload(registration_row):
    (
        registration_id,
        group_title,
        requester_name,
        child_name,
        phone_number,
        event_date,
    ) = registration_row

    payload = {
        "name": (requester_name or "").strip(),
        "phone": re.sub(r"\D", "", phone_number or ""),
        "date": (event_date or "").strip(),
        "rest": resolve_google_mk_restaurant(group_title),
        "children": str(infer_children_count(child_name)),
        "childrenNames": normalized_children_names(child_name),
    }
    validate_google_mk_payload(payload)
    return registration_id, payload


def mark_google_mk_sync(registration_id, *, synced, error=None):
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            INSERT INTO mk_google_sync(
                registration_id, synced, attempts, last_error, synced_at
            )
            VALUES(?,?,?,?,?)
            ON CONFLICT(registration_id) DO UPDATE SET
                synced=excluded.synced,
                attempts=mk_google_sync.attempts + 1,
                last_error=excluded.last_error,
                synced_at=excluded.synced_at
            """,
            (
                registration_id,
                1 if synced else 0,
                1,
                None if synced else (error or "Невідома помилка"),
                datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S") if synced else None,
            ),
        )
        conn.commit()


async def sync_mk_registration_to_google(registration_id):
    """Надсилає один локально збережений запис у Google Apps Script."""
    if not GOOGLE_MK_WEB_APP_URL:
        mark_google_mk_sync(
            registration_id,
            synced=False,
            error="GOOGLE_MK_WEB_APP_URL не заданий",
        )
        return False

    with sqlite3.connect(DATABASE) as conn:
        row = conn.execute(
            """
            SELECT id, group_title, requester_name, child_name,
                   phone_number, event_date
            FROM mk_registrations
            WHERE id=?
            """,
            (registration_id,),
        ).fetchone()

    if not row:
        return False

    try:
        _, payload = build_google_mk_payload(row)

        timeout = httpx.Timeout(8.0, connect=4.0)
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
            response = await client.get(GOOGLE_MK_WEB_APP_URL, params=payload)
            response.raise_for_status()

        try:
            result = response.json()
        except ValueError as exc:
            raise RuntimeError("Google Apps Script повернув не JSON") from exc

        if result.get("response") != "ok" or int(result.get("code", 0)) != 200:
            raise RuntimeError(f"Некоректна відповідь Google Apps Script: {result}")

        mark_google_mk_sync(registration_id, synced=True)
        logging.info(
            "Google MK: запис %s синхронізовано (%s, %s)",
            registration_id,
            payload["rest"],
            payload["date"],
        )
        return True

    except Exception as exc:
        mark_google_mk_sync(registration_id, synced=False, error=str(exc))
        logging.exception(
            "Google MK: не вдалося синхронізувати запис %s",
            registration_id,
        )
        return False


async def retry_pending_google_mk_sync(application):
    """Періодично повторює синхронізацію записів, які не вдалося відправити."""
    while True:
        await asyncio.sleep(300)

        try:
            with sqlite3.connect(DATABASE) as conn:
                rows = conn.execute(
                    """
                    SELECT r.id
                    FROM mk_registrations r
                    LEFT JOIN mk_google_sync s ON s.registration_id=r.id
                    WHERE r.event_date IS NOT NULL
                      AND TRIM(COALESCE(r.event_date, '')) <> ''
                      AND (s.registration_id IS NULL OR s.synced=0)
                      AND COALESCE(s.attempts, 0) < 20
                    ORDER BY r.id
                    LIMIT 50
                    """
                ).fetchall()

            for (registration_id,) in rows:
                await sync_mk_registration_to_google(registration_id)
                await asyncio.sleep(0.25)

        except asyncio.CancelledError:
            raise
        except Exception:
            logging.exception("Помилка фонового повтору Google MK sync")


async def _complete_mk_registration_with_schedule(
    *, context, group_id, group_title, user_id, requester_name,
    child_name, phone_number, event_date, event_title,
    source_message_id=None, phone_was_in_group=False, prompt_message_id=None
):
    created_at = datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
    phone_number = normalize_phone(phone_number)
    save_mk_phone(user_id, phone_number)

    with sqlite3.connect(DATABASE) as conn:
        cursor = conn.execute(
            """
            INSERT INTO mk_registrations(
                group_id, group_title, requester_id, requester_name,
                child_name, phone_number, created_at, event_date, event_title
            )
            VALUES(?,?,?,?,?,?,?,?,?)
            """,
            (
                group_id, group_title, user_id, requester_name,
                child_name, phone_number, created_at, event_date, event_title,
            ),
        )
        registration_id = cursor.lastrowid
        conn.commit()

    # Локальний Excel не повинен затримувати відповідь користувачу.
    # Оновлюємо його у фоновому потоці, щоб openpyxl не блокував Telegram-бота.
    try:
        context.application.create_task(
            asyncio.to_thread(refresh_mk_registrations_excel),
            name=f"refresh-mk-excel-{registration_id}",
        )
    except Exception:
        logging.exception(
            "Не вдалося запустити фонове оновлення Excel для запису %s",
            registration_id,
        )

    # Google Sheets також синхронізуємо у фоні.
    # Запис уже надійно збережений у SQLite, тому користувач не чекає
    # відповіді Google Apps Script і одразу отримує підтвердження.
    try:
        context.application.create_task(
            sync_mk_registration_to_google(registration_id),
            name=f"google-mk-sync-{registration_id}",
        )
    except Exception:
        logging.exception(
            "Не вдалося запустити Google sync для запису %s",
            registration_id,
        )

    _, confirmation_text = get_mk_settings(group_id)
    confirmation_text = format_mk_html(
        confirmation_text, child_name, requester_name, group_title or ""
    )
    event_display = datetime.strptime(event_date, "%Y-%m-%d").strftime("%d.%m.%Y")
    confirmation_text += (
        f"\n📅 {html.escape(event_display, quote=False)} — "
        f"{html.escape(str(event_title or ''), quote=False)}"
    )

    try:
        await context.bot.send_message(
            chat_id=group_id,
            text=confirmation_text,
            parse_mode="HTML",
            reply_to_message_id=source_message_id,
            allow_sending_without_reply=True,
        )
    except Exception:
        logging.exception("Не вдалося надіслати підтвердження в групу %s", group_id)

    # Після появи підсумкового тексту прибираємо службове повідомлення з кнопкою.
    await safe_delete_message(context, group_id, prompt_message_id)



async def _create_mk_request(
    message, chat, user, child_name, context, phone_number=None
):
    """Спочатку пропонує 4–5 дат, а вже потім завершує запис/просить номер."""
    register_group(chat)
    schedule = get_active_mk_schedule(chat.id)
    if not schedule:
        temp_message = await message.reply_text(
            "На жаль, актуальний графік майстер-класів ще не додано. "
            "Адміністратор опублікує його найближчим часом."
        )
        schedule_message_deletion(context, chat.id, temp_message.message_id)
        return

    token = secrets.token_urlsafe(12)
    user_name = user.full_name or str(user.id)
    explicit_phone = bool(phone_number)
    phone_number = phone_number or get_saved_mk_phone(user.id)

    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            INSERT INTO mk_booking_drafts(
                token, group_id, group_title, requester_id, requester_name,
                child_name, phone_number, source_message_id,
                phone_was_in_group, created_at
            )
            VALUES(?,?,?,?,?,?,?,?,?,?)
            """,
            (
                token, chat.id, chat.title, user.id, user_name,
                child_name, phone_number, message.message_id,
                1 if explicit_phone else 0,
                datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        conn.commit()

    buttons = []
    for schedule_id, date_text, title in schedule:
        date_label = datetime.strptime(date_text, "%Y-%m-%d").strftime("%d.%m")
        buttons.append(
            [
                InlineKeyboardButton(
                    f"📅 {date_label} — {title}",
                    callback_data=f"mkdate:{token}:{schedule_id}",
                )
            ]
        )

    selection_message = await message.reply_text(
        f"Оберіть майстер-клас для «{child_name}»:",
        reply_markup=InlineKeyboardMarkup(buttons),
    )

    # Повідомлення гостя залишаємо в групі. Видаляються лише службові
    # повідомлення самого бота, коли вони більше не потрібні.


async def mk_date_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    try:
        _, token, schedule_id_text = query.data.split(":", 2)
        schedule_id = int(schedule_id_text)
    except (ValueError, AttributeError):
        await query.answer("Некоректний вибір.", show_alert=True)
        return

    with sqlite3.connect(DATABASE) as conn:
        draft = conn.execute(
            """
            SELECT group_id, group_title, requester_id, requester_name,
                   child_name, phone_number, source_message_id, phone_was_in_group
            FROM mk_booking_drafts WHERE token=?
            """,
            (token,),
        ).fetchone()
        schedule = conn.execute(
            "SELECT event_date, title, group_id FROM mk_schedule WHERE id=?",
            (schedule_id,),
        ).fetchone()

    if not draft or not schedule:
        try:
            await query.edit_message_text("❌ Цей вибір уже недійсний.")
        except Exception:
            try:
                await query.answer("❌ Цей вибір уже недійсний.", show_alert=True)
            except Exception:
                pass
        return

    (
        group_id, group_title, requester_id, requester_name,
        child_name, phone_number, source_message_id, phone_was_in_group,
    ) = draft
    event_date, event_title, schedule_group_id = schedule

    if requester_id != query.from_user.id:
        await query.answer("Ця кнопка призначена для іншого користувача.", show_alert=True)
        return
    if schedule_group_id != group_id:
        await query.edit_message_text("❌ Майстер-клас не належить цій групі.")
        return

    with sqlite3.connect(DATABASE) as conn:
        conn.execute("DELETE FROM mk_booking_drafts WHERE token=?", (token,))
        conn.commit()

    if phone_number:
        await _complete_mk_registration_with_schedule(
            context=context,
            group_id=group_id,
            group_title=group_title,
            user_id=requester_id,
            requester_name=requester_name,
            child_name=child_name,
            phone_number=phone_number,
            event_date=event_date,
            event_title=event_title,
            source_message_id=source_message_id,
            phone_was_in_group=bool(phone_was_in_group),
        )
        # Підсумковий налаштований текст уже надіслано в групу.
        # Повідомлення з кнопками більше не потрібне.
        await safe_delete_message(
            context,
            query.message.chat_id,
            query.message.message_id,
        )
        return

    pending_token = secrets.token_urlsafe(18)
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            INSERT INTO mk_pending(
                token, group_id, group_title, requester_id, requester_name,
                child_name, source_message_id, created_at, status,
                event_date, event_title, prompt_message_id
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                pending_token, group_id, group_title, requester_id, requester_name,
                child_name, source_message_id,
                datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S"),
                "waiting_start", event_date, event_title,
                query.message.message_id,
            ),
        )
        conn.commit()

    bot_info = await context.bot.get_me()
    private_url = f"https://t.me/{bot_info.username}?start=mk_{pending_token}"
    response_text, _ = get_mk_settings(group_id)
    response_text = format_mk_html(
        response_text, child_name, requester_name, group_title or ""
    )
    event_display = datetime.strptime(event_date, "%Y-%m-%d").strftime("%d.%m.%Y")
    response_text += (
        f"\n\n📅 {html.escape(event_display, quote=False)} — "
        f"{html.escape(str(event_title or ''), quote=False)}"
    )

    await query.edit_message_text(
        response_text,
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton("📱 Надіслати номер телефону", url=private_url)]]
        ),
    )


async def begin_mk_private_registration(update, context, token):
    message = update.effective_message
    user = update.effective_user
    if update.effective_chat.type != "private":
        return

    with sqlite3.connect(DATABASE) as conn:
        pending = conn.execute(
            """
            SELECT group_id, group_title, requester_id, requester_name,
                   child_name, status, event_date, event_title, prompt_message_id
            FROM mk_pending WHERE token=?
            """,
            (token,),
        ).fetchone()

    if not pending:
        await message.reply_text("❌ Посилання недійсне або запис уже завершено.")
        return

    (
        group_id, group_title, requester_id, requester_name,
        child_name, status, event_date, event_title, prompt_message_id,
    ) = pending

    if requester_id != user.id:
        await message.reply_text("❌ Це посилання створене для іншого користувача.")
        return
    if status == "completed":
        await message.reply_text("✅ Цей запис уже завершено.")
        return

    saved_phone = get_saved_mk_phone(user.id)
    if saved_phone:
        await _complete_mk_registration_with_schedule(
            context=context,
            group_id=group_id,
            group_title=group_title,
            user_id=user.id,
            requester_name=requester_name,
            child_name=child_name,
            phone_number=saved_phone,
            event_date=event_date,
            event_title=event_title,
            prompt_message_id=prompt_message_id,
        )
        with sqlite3.connect(DATABASE) as conn:
            conn.execute("UPDATE mk_pending SET status='completed' WHERE token=?", (token,))
            conn.commit()
        await message.reply_text(
            f"✅ Запис для «{child_name}» створено. Використано збережений номер."
        )
        return

    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            "UPDATE mk_pending SET status='waiting_contact' WHERE token=?",
            (token,),
        )
        conn.execute(
            """
            INSERT INTO mk_user_state(user_id, token)
            VALUES(?,?)
            ON CONFLICT(user_id) DO UPDATE SET token=excluded.token
            """,
            (user.id, token),
        )
        conn.commit()

    keyboard = ReplyKeyboardMarkup(
        [[KeyboardButton("📱 Надіслати свій номер", request_contact=True)]],
        resize_keyboard=True,
        one_time_keyboard=True,
        input_field_placeholder="Натисніть кнопку нижче",
    )
    event_display = datetime.strptime(event_date, "%Y-%m-%d").strftime("%d.%m.%Y")
    await message.reply_text(
        f"Запис для «{child_name}».\n"
        f"📅 {event_display} — {event_title}\n\n"
        "Натисніть кнопку та поділіться своїм номером.",
        reply_markup=keyboard,
    )


async def receive_mk_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.effective_message
    user = update.effective_user
    contact = message.contact if message else None

    if not message or not user or not contact or update.effective_chat.type != "private":
        return
    if contact.user_id is not None and contact.user_id != user.id:
        await message.reply_text("❌ Надішліть саме свій номер кнопкою нижче.")
        return

    with sqlite3.connect(DATABASE) as conn:
        state = conn.execute(
            "SELECT token FROM mk_user_state WHERE user_id=?", (user.id,)
        ).fetchone()
        if not state:
            await message.reply_text(
                "Немає активного запису.", reply_markup=ReplyKeyboardRemove()
            )
            return

        token = state[0]
        pending = conn.execute(
            """
            SELECT group_id, group_title, requester_name, child_name,
                   event_date, event_title, prompt_message_id
            FROM mk_pending
            WHERE token=? AND requester_id=? AND status='waiting_contact'
            """,
            (token, user.id),
        ).fetchone()

    if not pending:
        with sqlite3.connect(DATABASE) as conn:
            conn.execute("DELETE FROM mk_user_state WHERE user_id=?", (user.id,))
            conn.commit()
        await message.reply_text(
            "❌ Активний запис не знайдено.", reply_markup=ReplyKeyboardRemove()
        )
        return

    (
        group_id, group_title, requester_name, child_name,
        event_date, event_title, prompt_message_id,
    ) = pending
    phone_number = normalize_phone(contact.phone_number)

    await _complete_mk_registration_with_schedule(
        context=context,
        group_id=group_id,
        group_title=group_title,
        user_id=user.id,
        requester_name=requester_name,
        child_name=child_name,
        phone_number=phone_number,
        event_date=event_date,
        event_title=event_title,
        prompt_message_id=prompt_message_id,
    )

    with sqlite3.connect(DATABASE) as conn:
        conn.execute("UPDATE mk_pending SET status='completed' WHERE token=?", (token,))
        conn.execute("DELETE FROM mk_user_state WHERE user_id=?", (user.id,))
        conn.commit()

    try:
        await context.bot.delete_message(message.chat_id, message.message_id)
    except Exception:
        logging.exception("Не вдалося видалити повідомлення з контактом")

    await context.bot.send_message(
        chat_id=user.id,
        text=f"✅ Запис для «{child_name}» успішно створено.",
        reply_markup=ReplyKeyboardRemove(),
    )


async def mk_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return

    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(
            """
            SELECT group_title, event_date, event_title, child_name,
                   requester_name, phone_number, created_at
            FROM mk_registrations
            ORDER BY id DESC LIMIT 20
            """
        ).fetchall()

    if not rows:
        await update.message.reply_text("Записів на майстер-класи ще немає.")
        return

    parts = ["📝 Останні записи на МК:\n"]
    for group_title, event_date, event_title, child_name, parent_name, phone, created_at in rows:
        event_display = event_date or "Не вказано"
        if event_date:
            try:
                event_display = datetime.strptime(event_date, "%Y-%m-%d").strftime("%d.%m.%Y")
            except ValueError:
                pass
        parts.append(
            f"\n🍣 {group_title}\n"
            f"Майстер-клас: {event_title or 'Не вказано'}\n"
            f"Дата МК: {event_display}\n"
            f"Ім’я дитини: {child_name}\n"
            f"Ім’я батьків: {parent_name}\n"
            f"Телефон: {phone}\n"
            f"Дата запису: {created_at}\n"
        )

    text = "".join(parts)
    for start_index in range(0, len(text), 4000):
        await update.message.reply_text(text[start_index:start_index + 4000])


async def mk_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    if update.effective_chat.type != "private":
        return

    filepath = build_mk_registrations_excel()
    with filepath.open("rb") as file:
        await update.message.reply_document(
            document=file,
            filename=filepath.name,
            caption="📝 Постійна таблиця записів на майстер-класи",
        )


async def cancel_mk_setting(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    mk_config_state.pop(update.effective_user.id, None)
    mk_schedule_state.pop(update.effective_user.id, None)
    await update.message.reply_text("❌ Налаштування МК скасовано.")



# ============================================================
# СИСТЕМА РОЛЕЙ І ДОСТУПУ
# ============================================================

access_state = {}

ROLE_DEVELOPER = "developer"
ROLE_MANAGER = "manager"
ROLE_ADMIN = "admin"

ROLE_NAMES = {
    ROLE_DEVELOPER: "👑 Розробник",
    ROLE_MANAGER: "👔 Керуючий",
    ROLE_ADMIN: "👨‍💼 Адміністратор",
}


def init_access_schema():
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS access_roles(
                user_id INTEGER PRIMARY KEY,
                role TEXT NOT NULL,
                display_name TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS admin_groups(
                user_id INTEGER NOT NULL,
                group_id INTEGER NOT NULL,
                PRIMARY KEY(user_id, group_id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bot_users(
                user_id INTEGER PRIMARY KEY,
                username TEXT,
                full_name TEXT,
                custom_name TEXT,
                first_start_at TEXT NOT NULL,
                last_start_at TEXT NOT NULL
            )
            """
        )
        now = datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
        conn.execute(
            """
            INSERT INTO access_roles(user_id, role, display_name, created_at)
            VALUES(?,?,?,?)
            ON CONFLICT(user_id) DO UPDATE SET role=excluded.role
            """,
            (DEVELOPER_ID, ROLE_DEVELOPER, "Розробник", now),
        )
        conn.execute(
            """
            INSERT INTO access_roles(user_id, role, display_name, created_at)
            VALUES(?,?,?,?)
            ON CONFLICT(user_id) DO UPDATE SET role=excluded.role
            """,
            (DEFAULT_MANAGER_ID, ROLE_MANAGER, "Керуючий", now),
        )
        conn.commit()


def get_user_role(user_id):
    if user_id == DEVELOPER_ID:
        return ROLE_DEVELOPER
    with sqlite3.connect(DATABASE) as conn:
        row = conn.execute(
            "SELECT role FROM access_roles WHERE user_id=?",
            (user_id,),
        ).fetchone()
    return row[0] if row else None


def is_admin(user_id):
    """Глобальний доступ: розробник або керуючий."""
    return get_user_role(user_id) in {ROLE_DEVELOPER, ROLE_MANAGER}


def can_manage_mk(user_id):
    return get_user_role(user_id) in {ROLE_DEVELOPER, ROLE_MANAGER, ROLE_ADMIN}


def can_manage_roles(user_id):
    return get_user_role(user_id) in {ROLE_DEVELOPER, ROLE_MANAGER}


def can_access_group(user_id, group_id):
    role = get_user_role(user_id)
    if role in {ROLE_DEVELOPER, ROLE_MANAGER}:
        return True
    if role != ROLE_ADMIN:
        return False
    with sqlite3.connect(DATABASE) as conn:
        row = conn.execute(
            "SELECT 1 FROM admin_groups WHERE user_id=? AND group_id=?",
            (user_id, group_id),
        ).fetchone()
    return bool(row)


def get_accessible_groups(user_id):
    role = get_user_role(user_id)
    if role in {ROLE_DEVELOPER, ROLE_MANAGER}:
        return get_groups()
    if role != ROLE_ADMIN:
        return []
    with sqlite3.connect(DATABASE) as conn:
        return conn.execute(
            """
            SELECT g.group_id, g.title
            FROM groups g
            JOIN admin_groups ag ON ag.group_id=g.group_id
            WHERE ag.user_id=?
            ORDER BY g.title
            """,
            (user_id,),
        ).fetchall()


def upsert_role(user_id, role, display_name=None):
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            INSERT INTO access_roles(user_id, role, display_name, created_at)
            VALUES(?,?,?,?)
            ON CONFLICT(user_id) DO UPDATE SET
                role=excluded.role,
                display_name=COALESCE(excluded.display_name, access_roles.display_name)
            """,
            (
                user_id,
                role,
                display_name,
                datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        conn.commit()


async def add_manager(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != DEVELOPER_ID:
        await update.message.reply_text("❌ Лише розробник може додавати керуючих.")
        return
    if update.effective_chat.type != "private":
        return
    if context.args:
        try:
            target_id = int(context.args[0])
        except ValueError:
            await update.message.reply_text("❌ Telegram ID має складатися з цифр.")
            return
        upsert_role(target_id, ROLE_MANAGER)
        await update.message.reply_text(f"✅ Керуючого додано: {target_id}")
        return
    user_id = update.effective_user.id
    mk_schedule_state.pop(user_id, None)
    mk_config_state.pop(user_id, None)
    broadcast_wait.pop(user_id, None)
    access_state[user_id] = {"action": "add_manager", "step": "waiting_id"}
    await update.message.reply_text("Введіть Telegram ID нового керуючого.")


async def add_admin_role(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not can_manage_roles(update.effective_user.id):
        return
    if update.effective_chat.type != "private":
        return
    user_id = update.effective_user.id
    mk_schedule_state.pop(user_id, None)
    mk_config_state.pop(user_id, None)
    broadcast_wait.pop(user_id, None)

    state = {"action": "add_admin", "step": "waiting_id"}
    access_state[user_id] = state
    if context.args:
        try:
            state["target_id"] = int(context.args[0])
        except ValueError:
            await update.message.reply_text("❌ Telegram ID має складатися з цифр.")
            return
        await show_admin_group_selection(update.effective_user.id, update.message, state["target_id"])
        state["step"] = "waiting_group"
    else:
        await update.message.reply_text("Введіть Telegram ID адміністратора.")


async def show_admin_group_selection(owner_id, message, target_id):
    groups_list = get_groups()
    if not groups_list:
        access_state.pop(owner_id, None)
        await message.reply_text("❌ Немає збережених груп.")
        return
    keyboard = [
        [InlineKeyboardButton(
            title if len(title) <= 48 else title[:45] + "...",
            callback_data=f"rolegrp:{target_id}:{group_id}",
        )]
        for group_id, title in groups_list
    ]
    keyboard.append([InlineKeyboardButton("✅ Завершити", callback_data=f"rolegrp:{target_id}:done")])
    keyboard.append([InlineKeyboardButton("❌ Скасувати", callback_data=f"rolegrp:{target_id}:cancel")])
    await message.reply_text(
        "До якої групи надати доступ? Можна вибрати декілька груп, потім натиснути «Завершити».",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


async def receive_access_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    message = update.effective_message
    if not user or not message or not message.text or update.effective_chat.type != "private":
        return
    state = access_state.get(user.id)
    if not state or state.get("step") != "waiting_id":
        return
    try:
        target_id = int(message.text.strip())
    except ValueError:
        await message.reply_text("❌ Telegram ID має складатися лише з цифр.")
        return

    # При активному сценарії доступу прибираємо старі сценарії МК.
    mk_schedule_state.pop(user.id, None)
    mk_config_state.pop(user.id, None)

    if state["action"] == "add_manager":
        if user.id != DEVELOPER_ID:
            access_state.pop(user.id, None)
            return
        upsert_role(target_id, ROLE_MANAGER)
        access_state.pop(user.id, None)
        await message.reply_text(f"✅ Керуючого додано: {target_id}")
        return

    if state["action"] == "add_admin":
        state["target_id"] = target_id
        state["step"] = "waiting_group"
        await show_admin_group_selection(user.id, message, target_id)


async def role_group_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    actor_id = query.from_user.id
    if not can_manage_roles(actor_id):
        await query.answer("Немає доступу.", show_alert=True)
        return
    try:
        _, target_text, group_text = query.data.split(":", 2)
        target_id = int(target_text)
    except (ValueError, AttributeError):
        await query.edit_message_text("❌ Некоректні дані.")
        return

    if group_text == "cancel":
        access_state.pop(actor_id, None)
        await query.edit_message_text("❌ Додавання адміністратора скасовано.")
        return

    if group_text == "done":
        with sqlite3.connect(DATABASE) as conn:
            count = conn.execute(
                "SELECT COUNT(*) FROM admin_groups WHERE user_id=?",
                (target_id,),
            ).fetchone()[0]
        if count == 0:
            await query.answer("Спочатку виберіть хоча б одну групу.", show_alert=True)
            return
        upsert_role(target_id, ROLE_ADMIN)
        access_state.pop(actor_id, None)
        await query.edit_message_text(
            f"✅ Адміністратора додано.\nTelegram ID: {target_id}\nГруп призначено: {count}"
        )
        return

    try:
        group_id = int(group_text)
    except ValueError:
        return
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            "INSERT OR IGNORE INTO admin_groups(user_id, group_id) VALUES(?,?)",
            (target_id, group_id),
        )
        conn.commit()
    await query.answer("Групу додано. Можна вибрати ще одну.", show_alert=False)


async def delete_manager(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != DEVELOPER_ID:
        await update.message.reply_text("❌ Лише розробник може видаляти керуючих.")
        return
    if not context.args:
        await update.message.reply_text("Використання: /delmanager TELEGRAM_ID")
        return
    try:
        target_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Некоректний Telegram ID.")
        return
    if target_id in {DEVELOPER_ID, DEFAULT_MANAGER_ID}:
        await update.message.reply_text("❌ Базову роль видалити не можна.")
        return
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("DELETE FROM access_roles WHERE user_id=? AND role=?", (target_id, ROLE_MANAGER))
        conn.commit()
    await update.message.reply_text(f"✅ Керуючого видалено: {target_id}")


async def delete_admin_role(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not can_manage_roles(update.effective_user.id):
        return
    if not context.args:
        await update.message.reply_text("Використання: /deladmin TELEGRAM_ID")
        return
    try:
        target_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Некоректний Telegram ID.")
        return
    if get_user_role(target_id) in {ROLE_DEVELOPER, ROLE_MANAGER}:
        await update.message.reply_text("❌ Не можна видалити користувача з вищою роллю.")
        return
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("DELETE FROM admin_groups WHERE user_id=?", (target_id,))
        conn.execute("DELETE FROM access_roles WHERE user_id=? AND role=?", (target_id, ROLE_ADMIN))
        conn.commit()
    await update.message.reply_text(f"✅ Адміністратора видалено: {target_id}")


async def roles_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not can_manage_roles(update.effective_user.id):
        return
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(
            "SELECT user_id, role, COALESCE(display_name, '') FROM access_roles ORDER BY role, user_id"
        ).fetchall()
        group_rows = conn.execute(
            """
            SELECT ag.user_id, g.title
            FROM admin_groups ag
            LEFT JOIN groups g ON g.group_id=ag.group_id
            ORDER BY ag.user_id, g.title
            """
        ).fetchall()
    groups_by_user = {}
    for user_id, title in group_rows:
        groups_by_user.setdefault(user_id, []).append(title or "Невідома група")
    sections = []
    for role in (ROLE_DEVELOPER, ROLE_MANAGER, ROLE_ADMIN):
        sections.append(ROLE_NAMES[role])
        role_rows = [row for row in rows if row[1] == role]
        if not role_rows:
            sections.append("— немає")
        for user_id, _, display_name in role_rows:
            label = f"{display_name + ' — ' if display_name else ''}{user_id}"
            sections.append(label)
            for title in groups_by_user.get(user_id, []):
                sections.append(f"  • {title}")
        sections.append("")
    await update.message.reply_text("\n".join(sections)[:4000])


async def admins_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not can_manage_roles(update.effective_user.id):
        return
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(
            """
            SELECT r.user_id, COALESCE(r.display_name, ''), g.title
            FROM access_roles r
            LEFT JOIN admin_groups ag ON ag.user_id=r.user_id
            LEFT JOIN groups g ON g.group_id=ag.group_id
            WHERE r.role=?
            ORDER BY r.user_id, g.title
            """,
            (ROLE_ADMIN,),
        ).fetchall()
    if not rows:
        await update.message.reply_text("Адміністраторів ще немає.")
        return
    data = {}
    for user_id, name, title in rows:
        entry = data.setdefault(user_id, {"name": name, "groups": []})
        if title:
            entry["groups"].append(title)
    lines = ["👨‍💼 Адміністратори\n"]
    for user_id, info in data.items():
        lines.append(f"{info['name'] + ' — ' if info['name'] else ''}{user_id}")
        for title in info["groups"]:
            lines.append(f"  • {title}")
        lines.append("")
    await update.message.reply_text("\n".join(lines)[:4000])


def build_mk_registrations_excel_for_user(user_id):
    role = get_user_role(user_id)

    if role in {ROLE_DEVELOPER, ROLE_MANAGER}:
        return build_mk_registrations_excel()

    allowed_groups = get_accessible_groups(user_id)
    allowed_ids = [group_id for group_id, _ in allowed_groups]

    if not allowed_ids:
        return None

    placeholders = ",".join("?" for _ in allowed_ids)

    with _mk_excel_lock:
        with sqlite3.connect(DATABASE) as conn:
            rows = conn.execute(
                f"""
                SELECT group_id, group_title, created_at, event_date, event_title,
                       child_name, requester_name, phone_number
                FROM mk_registrations
                WHERE group_id IN ({placeholders})
                ORDER BY COALESCE(created_at, '') DESC, id DESC
                """,
                allowed_ids,
            ).fetchall()
            schedules = conn.execute(
                f"""
                SELECT group_id, event_date, title
                FROM mk_schedule
                WHERE group_id IN ({placeholders})
                ORDER BY event_date, id
                """,
                allowed_ids,
            ).fetchall()

        filepath = Path(f"Записи_МК_{user_id}.xlsx")
        return _build_mk_workbook(rows, allowed_groups, schedules, filepath)


async def mk_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Відкриває сайт майстер-класів у Telegram Mini App."""
    user_id = update.effective_user.id

    if not can_manage_mk(user_id) or update.effective_chat.type != "private":
        return

    if not WEB_APP_URL:
        await update.message.reply_text(
            "❌ WEB_APP_URL не налаштовано. Додайте HTTPS-адресу сайту в змінні середовища."
        )
        return

    keyboard = InlineKeyboardMarkup(
        [[
            InlineKeyboardButton(
                "🍣 Відкрити майстер-класи",
                web_app=WebAppInfo(url=WEB_APP_URL),
            )
        ]]
    )

    await update.message.reply_text(
        "📝 Майстер-класи\n\n"
        "Сайт працює з тією самою базою даних, що й бот.",
        reply_markup=keyboard,
    )


async def mk_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not can_manage_mk(user_id):
        return
    allowed = [group_id for group_id, _ in get_accessible_groups(user_id)]
    role = get_user_role(user_id)
    with sqlite3.connect(DATABASE) as conn:
        if role in {ROLE_DEVELOPER, ROLE_MANAGER}:
            rows = conn.execute(
                """
                SELECT group_title, event_date, event_title, child_name,
                       requester_name, phone_number, created_at
                FROM mk_registrations ORDER BY id DESC LIMIT 20
                """
            ).fetchall()
        elif allowed:
            placeholders = ",".join("?" for _ in allowed)
            rows = conn.execute(
                f"""
                SELECT group_title, event_date, event_title, child_name,
                       requester_name, phone_number, created_at
                FROM mk_registrations
                WHERE group_id IN ({placeholders})
                ORDER BY id DESC LIMIT 20
                """,
                allowed,
            ).fetchall()
        else:
            rows = []
    if not rows:
        await update.message.reply_text("Записів на майстер-класи ще немає.")
        return
    parts = ["📝 Останні записи на МК:\n"]
    for group_title, event_date, event_title, child_name, parent_name, phone, created_at in rows:
        parts.append(
            f"\n🍣 {group_title}\nДата МК: {event_date or 'Не вказано'}\n"
            f"Майстер-клас: {event_title or 'Не вказано'}\n"
            f"Ім’я дитини: {child_name}\nІм’я батьків: {parent_name}\n"
            f"Телефон: {phone}\nДата запису: {created_at}\n"
        )
    text = "".join(parts)
    for start_index in range(0, len(text), 4000):
        await update.message.reply_text(text[start_index:start_index + 4000])


async def cancel_mk_setting(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not can_manage_mk(update.effective_user.id):
        return
    mk_config_state.pop(update.effective_user.id, None)
    mk_schedule_state.pop(update.effective_user.id, None)
    access_state.pop(update.effective_user.id, None)
    await update.message.reply_text("❌ Поточне налаштування скасовано.")


def get_private_commands_for_role(role):
    common_mk = [
        BotCommand("setmkresponse", "Налаштувати відповідь для МК"),
        BotCommand("setmkconfirm", "Налаштувати підтвердження МК"),
        BotCommand("setmkschedule", "Додати або змінити графік МК"),
        BotCommand("mklist", "Показати записи МК"),
        BotCommand("mkexcel", "Відкрити сайт МК"),
        BotCommand("cancelmk", "Скасувати поточне налаштування"),
    ]

    if role == ROLE_DEVELOPER:
        return [
            BotCommand("start", "Відкрити меню бота"),
            BotCommand("groups", "Показати список груп"),
            BotCommand("stats", "Показати статистику груп"),
            BotCommand("excel", "Отримати Excel-звіт статистики"),
            BotCommand("broadcast", "Створити розсилку"),
            BotCommand("deletebroadcast", "Видалити останню розсилку"),
            BotCommand("delete_id", "Видалити повідомлення за ID"),
            *common_mk,
            BotCommand("addmanager", "Додати керуючого"),
            BotCommand("delmanager", "Видалити керуючого"),
            BotCommand("addadmin", "Додати адміністратора"),
            BotCommand("deladmin", "Видалити адміністратора"),
            BotCommand("admins", "Список адміністраторів"),
            BotCommand("roles", "Список усіх ролей"),
            BotCommand("setname", "Назвати Telegram ID"),
        ]

    if role == ROLE_MANAGER:
        return [
            BotCommand("start", "Відкрити меню бота"),
            BotCommand("groups", "Показати список груп"),
            BotCommand("stats", "Показати статистику груп"),
            BotCommand("excel", "Отримати Excel-звіт статистики"),
            BotCommand("broadcast", "Створити розсилку"),
            BotCommand("deletebroadcast", "Видалити останню розсилку"),
            BotCommand("delete_id", "Видалити повідомлення за ID"),
            *common_mk,
            BotCommand("addadmin", "Додати адміністратора"),
            BotCommand("deladmin", "Видалити адміністратора"),
            BotCommand("admins", "Список адміністраторів"),
            BotCommand("roles", "Список усіх ролей"),
            BotCommand("setname", "Назвати Telegram ID"),
        ]

    if role == ROLE_ADMIN:
        return [
            BotCommand("start", "Відкрити меню бота"),
            *common_mk,
        ]

    return [BotCommand("start", "Запустити бота")]


async def set_private_commands_for_user(bot, user_id):
    commands = get_private_commands_for_role(get_user_role(user_id))
    await bot.set_my_commands(
        commands,
        scope=BotCommandScopeChat(chat_id=user_id),
    )


def save_started_user(user):
    now = datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
    username = user.username or ""
    full_name = user.full_name or str(user.id)
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            INSERT INTO bot_users(
                user_id, username, full_name, custom_name,
                first_start_at, last_start_at
            )
            VALUES(?,?,?,?,?,?)
            ON CONFLICT(user_id) DO UPDATE SET
                username=excluded.username,
                full_name=excluded.full_name,
                last_start_at=excluded.last_start_at
            """,
            (user.id, username, full_name, None, now, now),
        )
        conn.commit()
    return now


def get_custom_user_name(user_id):
    with sqlite3.connect(DATABASE) as conn:
        row = conn.execute(
            """
            SELECT COALESCE(NULLIF(custom_name, ''), NULLIF(full_name, ''), '')
            FROM bot_users
            WHERE user_id=?
            """,
            (user_id,),
        ).fetchone()
    return row[0] if row and row[0] else ""


def get_start_notification_recipients():
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(
            "SELECT user_id FROM access_roles WHERE role IN (?,?)",
            (ROLE_DEVELOPER, ROLE_MANAGER),
        ).fetchall()
    recipients = {DEVELOPER_ID, DEFAULT_MANAGER_ID}
    recipients.update(row[0] for row in rows)
    return sorted(recipients)


async def notify_start_pressed(context, user, started_at, payload=None):
    username = f"@{user.username}" if user.username else "не вказаний"
    role = get_user_role(user.id)
    role_name = ROLE_NAMES.get(role, "Без ролі")
    custom_name = get_custom_user_name(user.id)
    text = (
        "▶️ Користувач натиснув /start\n\n"
        f"ID: {user.id}\n"
        f"Нік: {username}\n"
        f"Ім’я: {user.full_name or 'не вказано'}\n"
        f"Ваша назва: {custom_name or 'не задана'}\n"
        f"Роль: {role_name}\n"
        f"Час: {started_at}"
    )
    if payload:
        text += f"\nПараметр запуску: {payload}"

    for recipient_id in get_start_notification_recipients():
        try:
            await context.bot.send_message(chat_id=recipient_id, text=text)
        except Exception:
            logging.exception(
                "Не вдалося повідомити про /start користувачу %s",
                recipient_id,
            )


async def set_user_custom_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    actor_id = update.effective_user.id
    if get_user_role(actor_id) not in {ROLE_DEVELOPER, ROLE_MANAGER}:
        await update.message.reply_text("❌ У вас немає доступу до цієї команди.")
        return
    if update.effective_chat.type != "private":
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Використання:\n/setname TELEGRAM_ID НАЗВА\n\n"
            "Приклад:\n/setname 123456789 Адміністратор Retroville"
        )
        return
    try:
        target_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Telegram ID має складатися з цифр.")
        return

    custom_name = " ".join(context.args[1:]).strip()
    if not custom_name:
        await update.message.reply_text("❌ Назва не може бути порожньою.")
        return

    now = datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """
            INSERT INTO bot_users(
                user_id, username, full_name, custom_name,
                first_start_at, last_start_at
            )
            VALUES(?,?,?,?,?,?)
            ON CONFLICT(user_id) DO UPDATE SET
                custom_name=excluded.custom_name
            """,
            (target_id, "", "", custom_name, now, now),
        )
        conn.execute(
            "UPDATE access_roles SET display_name=? WHERE user_id=?",
            (custom_name, target_id),
        )
        conn.commit()

    await update.message.reply_text(
        f"✅ Для ID {target_id} встановлено назву:\n{custom_name}"
    )


def format_role_commands_text(role):
    commands = get_private_commands_for_role(role)
    return "\n".join(
        f"/{command.command} — {command.description}"
        for command in commands
    )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    message = update.effective_message

    if not user or not message:
        return

    # Зберігаємо користувача в базі, але більше не надсилаємо
    # Розробнику та Керуючим повідомлення при кожному /start.
    save_started_user(user)

    payload = context.args[0] if context.args else None
    await set_private_commands_for_user(context.bot, user.id)

    # Посилання запису на МК обробляється без повідомлення про роль.
    if payload and payload.startswith("mk_"):
        token = payload[3:]
        await begin_mk_private_registration(update, context, token)
        return

    role = get_user_role(user.id)
    commands_text = format_role_commands_text(role)

    if role == ROLE_DEVELOPER:
        text = (
            "👑 Ви увійшли як Розробник. "
            "Вам доступні всі функції бота.\n\n"
            "📋 Список команд:\n"
            f"{commands_text}"
        )

    elif role == ROLE_MANAGER:
        text = (
            "👔 Ви увійшли як Керуючий. "
            "Вам доступні всі робочі функції та керування адміністраторами.\n\n"
            "📋 Список команд:\n"
            f"{commands_text}"
        )

    elif role == ROLE_ADMIN:
        groups_list = get_accessible_groups(user.id)
        group_names = (
            "\n".join(f"• {title}" for _, title in groups_list)
            or "• групу ще не призначено"
        )

        text = (
            "👨‍💼 Ви увійшли як Адміністратор МК.\n\n"
            "🏷 Ваші групи:\n"
            f"{group_names}\n\n"
            "📋 Список команд:\n"
            f"{commands_text}"
        )

    else:
        text = (
            "✅ Бот активовано.\n\n"
            "Для запису на майстер-клас поверніться до групи МК."
        )

    await message.reply_text(text)


async def post_init(application):
    # Мінімальне меню за замовчуванням для нових користувачів.
    await application.bot.set_my_commands(
        [BotCommand("start", "Запустити бота")],
        scope=BotCommandScopeAllPrivateChats(),
    )

    # У групах не показуємо приватні адміністративні команди.
    await application.bot.set_my_commands(
        [BotCommand("addgroup", "Додати поточну групу")],
        scope=BotCommandScopeAllGroupChats(),
    )

    # Окреме меню для кожного вже відомого користувача з роллю.
    with sqlite3.connect(DATABASE) as conn:
        role_users = conn.execute(
            "SELECT user_id FROM access_roles"
        ).fetchall()
    for (user_id,) in role_users:
        try:
            await set_private_commands_for_user(application.bot, user_id)
        except Exception:
            logging.exception("Не вдалося встановити меню команд для %s", user_id)

    application.create_task(monthly_report_scheduler(application))
    application.create_task(
        retry_pending_google_mk_sync(application),
        name="google-mk-sync-retry",
    )



# ============================================================
# ІМПОРТ ІСНУЮЧИХ ЗАПИСІВ ІЗ GOOGLE SHEETS ДЛЯ MINI APP
# ============================================================


def _google_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _google_key(value):
    value = _google_text(value).casefold()
    value = value.replace("’", "'").replace("ʼ", "'").replace("`", "'")
    value = re.sub(r"[^0-9a-zа-яіїєґ' ]+", " ", value, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", value).strip()


def _google_phone_key(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if digits.startswith("0") and len(digits) == 10:
        digits = "38" + digits
    return digits


def _google_date_iso(value):
    value = _google_text(value)
    if not value:
        return ""

    # ISO datetime / ISO date.
    iso_match = re.match(r"^(\d{4}-\d{2}-\d{2})", value)
    if iso_match:
        return iso_match.group(1)

    for fmt in (
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d.%m.%y",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass

    return ""


def _google_created_at(value, event_date):
    value = _google_text(value)
    if value:
        for fmt in (
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
        ):
            try:
                return datetime.strptime(value[:19], fmt).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass

    return f"{event_date} 00:00:00" if event_date else datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")


def _google_registration_columns(headers):
    """
    Повертає індекси колонок головного листа.
    Підтримує українські/російські/англійські заголовки та fallback по позиціях.
    """
    normalized = [_google_key(h).replace(" ", "") for h in headers]

    aliases = {
        "created": (
            "датачасреєстрації", "датареєстрації", "часреєстрації",
            "датавремярегистрации", "timestamp", "createdat",
        ),
        "restaurant": ("ресторан", "restaurant", "заклад"),
        "event_date": (
            "датамк", "датамайстеркласу", "датамайстеркласса",
            "датаmasterclass", "date",
        ),
        "parent": (
            "імябатьків", "імябатька", "батьки", "имяродителей",
            "родители", "name", "guestname",
        ),
        "phone": ("телефон", "номер телефону", "номертелефону", "phone"),
        "children_count": (
            "кількістьдітей", "кводітей", "количестводетей", "children",
        ),
        "children_names": (
            "іменадітей", "імядитини", "имена детей", "именадетей",
            "childrennames", "childname",
        ),
    }

    result = {}
    for key, variants in aliases.items():
        wanted = {_google_key(v).replace(" ", "") for v in variants}
        for index, header in enumerate(normalized):
            if header in wanted:
                result[key] = index
                break

    # Структура наданої таблиці:
    # дата/час, ресторан, дата МК, батьки, телефон, кількість, імена.
    fallback = {
        "created": 0,
        "restaurant": 1,
        "event_date": 2,
        "parent": 3,
        "phone": 4,
        "children_count": 5,
        "children_names": 6,
    }
    for key, index in fallback.items():
        if key not in result and index < len(headers):
            result[key] = index

    return result


def _google_cell(row, columns, key):
    index = columns.get(key)
    if index is None or index >= len(row):
        return ""
    return _google_text(row[index])


def _google_sheet_csv_rows():
    """Завантажує існуючий лист 'Всі реєстрації' як CSV."""
    if not GOOGLE_MK_SHEET_ID:
        return []

    url = (
        f"https://docs.google.com/spreadsheets/d/"
        f"{GOOGLE_MK_SHEET_ID}/gviz/tq"
    )

    timeout = httpx.Timeout(8.0, connect=4.0)
    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        response = client.get(
            url,
            params={
                "tqx": "out:csv",
                "sheet": GOOGLE_MK_REGISTRATIONS_SHEET,
            },
        )
        response.raise_for_status()

    content_type = (response.headers.get("content-type") or "").lower()
    body_start = response.text[:300].casefold()

    # Якщо Google повернув сторінку входу замість CSV.
    if "text/html" in content_type and (
        "accounts.google.com" in body_start
        or "<html" in body_start
    ):
        raise RuntimeError(
            "Google-таблиця недоступна для читання без авторизації. "
            "Надайте доступ 'Усі, хто має посилання — Перегляд' "
            "або окремий endpoint читання."
        )

    return list(csv.reader(StringIO(response.text)))


def _google_restaurant_group_map():
    result = {}
    for group_id, title in get_groups():
        restaurant = resolve_google_mk_restaurant(title)
        result[_google_key(restaurant)] = (group_id, title)
    return result


def sync_google_sheet_registrations_to_local(force=False):
    """
    Підтягує записи з уже існуючої Google-таблиці в локальну SQLite.
    Це дозволяє існуючому /api/registrations показувати і записи з бота,
    і записи, які були створені безпосередньо через таблицю/форму.

    При помилці Google сайт продовжує працювати з локальними записами.
    """
    global _google_mk_last_read_at

    now_ts = time.time()
    if (
        not force
        and _google_mk_last_read_at
        and now_ts - _google_mk_last_read_at < GOOGLE_MK_READ_CACHE_SECONDS
    ):
        return 0

    with _google_mk_read_lock:
        now_ts = time.time()
        if (
            not force
            and _google_mk_last_read_at
            and now_ts - _google_mk_last_read_at < GOOGLE_MK_READ_CACHE_SECONDS
        ):
            return 0

        try:
            csv_rows = _google_sheet_csv_rows()
        except Exception:
            logging.exception("Google MK: не вдалося прочитати таблицю реєстрацій")
            _google_mk_last_read_at = now_ts
            return 0

        if len(csv_rows) < 2:
            _google_mk_last_read_at = now_ts
            return 0

        headers = csv_rows[0]
        columns = _google_registration_columns(headers)
        restaurant_map = _google_restaurant_group_map()
        imported = 0

        with sqlite3.connect(DATABASE, timeout=10) as conn:
            conn.execute("PRAGMA busy_timeout=10000")

            for csv_row in csv_rows[1:]:
                restaurant = _google_cell(csv_row, columns, "restaurant")
                event_date = _google_date_iso(
                    _google_cell(csv_row, columns, "event_date")
                )

                if not restaurant or not event_date:
                    continue

                group_info = restaurant_map.get(_google_key(restaurant))
                if not group_info:
                    # Запис іншого ресторану, якого немає серед Telegram-груп бота.
                    continue

                group_id, group_title = group_info
                parent_name = _google_cell(csv_row, columns, "parent")
                phone_raw = _google_cell(csv_row, columns, "phone")
                phone_key = _google_phone_key(phone_raw)
                phone_number = normalize_phone(phone_raw) if phone_key else ""
                child_name = _google_cell(csv_row, columns, "children_names")

                # На сайті показуємо саме імена. Якщо в рядку вони не заповнені —
                # не вигадуємо їх із кількості дітей.
                if not child_name:
                    child_name = "—"

                created_at = _google_created_at(
                    _google_cell(csv_row, columns, "created"),
                    event_date,
                )

                source_text = "|".join(
                    [
                        _google_key(restaurant),
                        event_date,
                        phone_key,
                        _google_key(parent_name),
                        _google_key(child_name),
                    ]
                )
                source_key = hashlib.sha256(
                    source_text.encode("utf-8")
                ).hexdigest()

                existing_source = conn.execute(
                    """
                    SELECT id
                    FROM mk_registrations
                    WHERE google_source_key=?
                    """,
                    (source_key,),
                ).fetchone()
                if existing_source:
                    continue

                # Не дублюємо локальний запис бота, який уже відправився в Google.
                candidates = conn.execute(
                    """
                    SELECT id, requester_name, child_name, phone_number
                    FROM mk_registrations
                    WHERE group_id=? AND event_date=?
                    """,
                    (group_id, event_date),
                ).fetchall()

                matched_id = None
                for reg_id, local_parent, local_child, local_phone in candidates:
                    same_phone = (
                        phone_key
                        and _google_phone_key(local_phone) == phone_key
                    )
                    same_parent = (
                        _google_key(local_parent) == _google_key(parent_name)
                    )
                    same_child = (
                        _google_key(local_child) == _google_key(child_name)
                    )

                    # Телефон + ім'я дитини достатні для надійного збігу;
                    # якщо телефону немає — використовуємо батьків + дитину.
                    if (
                        (same_phone and same_child)
                        or (not phone_key and same_parent and same_child)
                    ):
                        matched_id = reg_id
                        break

                if matched_id is not None:
                    conn.execute(
                        """
                        UPDATE mk_registrations
                        SET google_source_key=?
                        WHERE id=? AND google_source_key IS NULL
                        """,
                        (source_key, matched_id),
                    )
                    # Не відправляємо цей рядок назад у Google повторно.
                    conn.execute(
                        """
                        INSERT INTO mk_google_sync(
                            registration_id, synced, attempts, last_error, synced_at
                        )
                        VALUES(?,1,0,NULL,?)
                        ON CONFLICT(registration_id) DO UPDATE SET
                            synced=1,
                            last_error=NULL,
                            synced_at=excluded.synced_at
                        """,
                        (
                            matched_id,
                            datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S"),
                        ),
                    )
                    continue

                schedule_row = conn.execute(
                    """
                    SELECT title
                    FROM mk_schedule
                    WHERE group_id=? AND event_date=?
                    ORDER BY id
                    LIMIT 1
                    """,
                    (group_id, event_date),
                ).fetchone()
                event_title = schedule_row[0] if schedule_row else ""

                cursor = conn.execute(
                    """
                    INSERT INTO mk_registrations(
                        group_id,
                        group_title,
                        requester_id,
                        requester_name,
                        child_name,
                        phone_number,
                        created_at,
                        event_date,
                        event_title,
                        attendance_status,
                        google_source_key
                    )
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        group_id,
                        group_title,
                        0,
                        parent_name or "—",
                        child_name,
                        phone_number,
                        created_at,
                        event_date,
                        event_title,
                        "yes",
                        source_key,
                    ),
                )
                registration_id = cursor.lastrowid

                # Це вже запис із Google — не треба відправляти його назад.
                conn.execute(
                    """
                    INSERT INTO mk_google_sync(
                        registration_id, synced, attempts, last_error, synced_at
                    )
                    VALUES(?,1,0,NULL,?)
                    ON CONFLICT(registration_id) DO UPDATE SET
                        synced=1,
                        last_error=NULL,
                        synced_at=excluded.synced_at
                    """,
                    (
                        registration_id,
                        datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )
                imported += 1

            conn.commit()

        _google_mk_last_read_at = time.time()

        if imported:
            logging.info(
                "Google MK: імпортовано %s нових записів для Mini App",
                imported,
            )

        return imported



# ============================================================
# TELEGRAM MINI APP / WEB API
# ============================================================

WEB_DIR = Path(__file__).resolve().parent / "web"
web_app = FastAPI(title="Євразія — Майстер-класи")

# Firebase Hosting працює на іншому домені, тому дозволяємо його origin для API.
web_app.add_middleware(
    CORSMiddleware,
    allow_origins=WEB_ALLOWED_ORIGINS,
    allow_credentials=False,
    allow_methods=["GET", "PATCH", "OPTIONS"],
    allow_headers=["Content-Type", "X-Telegram-Init-Data"],
)

if (WEB_DIR / "static").exists():
    web_app.mount(
        "/static",
        StaticFiles(directory=str(WEB_DIR / "static")),
        name="static",
    )


def _web_db():
    conn = sqlite3.connect(DATABASE, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout=10000")
    return conn


def _verify_telegram_init_data(init_data: str):
    """Перевіряє Telegram WebApp initData та повертає Telegram user."""
    if not init_data:
        if WEB_DEBUG_USER_ID:
            return {"id": int(WEB_DEBUG_USER_ID), "first_name": "Debug"}
        raise HTTPException(status_code=401, detail="Відкрийте сайт через Telegram-бота")

    if not TOKEN:
        raise HTTPException(status_code=500, detail="BOT_TOKEN не налаштовано")

    values = dict(parse_qsl(init_data, keep_blank_values=True))
    received_hash = values.pop("hash", None)
    if not received_hash:
        raise HTTPException(status_code=401, detail="Некоректні дані Telegram")

    data_check_string = "\n".join(
        f"{key}={values[key]}" for key in sorted(values)
    )
    secret_key = hmac.new(
        b"WebAppData",
        TOKEN.encode("utf-8"),
        hashlib.sha256,
    ).digest()
    calculated_hash = hmac.new(
        secret_key,
        data_check_string.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()

    if not hmac.compare_digest(calculated_hash, received_hash):
        raise HTTPException(status_code=401, detail="Підпис Telegram не пройшов перевірку")

    try:
        auth_date = int(values.get("auth_date", "0"))
    except ValueError:
        raise HTTPException(status_code=401, detail="Некоректна дата авторизації")

    if not auth_date or abs(int(time.time()) - auth_date) > 86400:
        raise HTTPException(status_code=401, detail="Сеанс Telegram застарів. Відкрийте Mini App знову")

    try:
        user = json.loads(values.get("user", "{}"))
    except json.JSONDecodeError:
        raise HTTPException(status_code=401, detail="Некоректні дані користувача")

    if not user.get("id"):
        raise HTTPException(status_code=401, detail="Telegram ID не знайдено")
    return user


def _web_current_user(request: Request):
    tg_user = _verify_telegram_init_data(
        request.headers.get("X-Telegram-Init-Data", "")
    )
    user_id = int(tg_user["id"])
    role = get_user_role(user_id)
    if role not in {ROLE_DEVELOPER, ROLE_MANAGER, ROLE_ADMIN}:
        raise HTTPException(status_code=403, detail="У вас немає доступу до майстер-класів")

    groups = get_accessible_groups(user_id)
    return {
        "id": user_id,
        "role": role,
        "telegram": tg_user,
        "groups": groups,
    }


def _web_allowed_group_ids(current_user):
    return [group_id for group_id, _title in current_user["groups"]]


def _web_children_count(child_names):
    try:
        return infer_children_count(child_names)
    except Exception:
        return 1


@web_app.get("/api/health")
def web_health():
    return {"ok": True, "service": "evrasia-masterclass-api"}


@web_app.get("/")
def web_index():
    index_path = WEB_DIR / "index.html"
    if not index_path.exists():
        raise HTTPException(status_code=500, detail="web/index.html не знайдено")
    return FileResponse(index_path)


@web_app.get("/config.js")
def web_config():
    """Віддає конфіг фронтенду, якщо index.html його підключає."""
    config_path = WEB_DIR / "config.js"
    if not config_path.exists():
        raise HTTPException(status_code=404, detail="web/config.js не знайдено")
    return FileResponse(config_path, media_type="application/javascript")


@web_app.get("/api/me")
def web_me(request: Request):
    current = _web_current_user(request)
    display_name = get_custom_user_name(current["id"]) or current["telegram"].get("first_name", "")
    return {
        "id": current["id"],
        "name": display_name,
        "role": current["role"],
        "restaurants": [
            {"id": group_id, "name": title}
            for group_id, title in current["groups"]
        ],
    }


@web_app.get("/api/dates")
def web_dates(request: Request):
    current = _web_current_user(request)
    allowed = _web_allowed_group_ids(current)
    if not allowed:
        return {"dates": []}

    placeholders = ",".join("?" for _ in allowed)
    with _web_db() as conn:
        rows = conn.execute(
            f"""
            SELECT DISTINCT event_date
            FROM (
                SELECT event_date FROM mk_schedule
                WHERE group_id IN ({placeholders})
                UNION
                SELECT event_date FROM mk_registrations
                WHERE group_id IN ({placeholders})
            )
            WHERE event_date IS NOT NULL AND event_date <> ''
            ORDER BY event_date
            """,
            [*allowed, *allowed],
        ).fetchall()
    return {"dates": [row[0] for row in rows]}


@web_app.get("/api/restaurants")
def web_restaurants(request: Request, date: str | None = None):
    current = _web_current_user(request)

    # Додаємо в локальне представлення нові рядки з Google-таблиці.
    sync_google_sheet_registrations_to_local()

    result = []
    with _web_db() as conn:
        for group_id, title in current["groups"]:
            if date:
                count = conn.execute(
                    "SELECT COUNT(*) FROM mk_registrations WHERE group_id=? AND event_date=?",
                    (group_id, date),
                ).fetchone()[0]
            else:
                count = conn.execute(
                    "SELECT COUNT(*) FROM mk_registrations WHERE group_id=?",
                    (group_id,),
                ).fetchone()[0]
            result.append({"id": group_id, "name": title, "count": count})
    return {"restaurants": result}


@web_app.get("/api/registrations")
def web_registrations(
    request: Request,
    date: str,
    group_id: int | None = None,
    q: str = "",
):
    current = _web_current_user(request)

    # Підтягуємо записи, які існують у Google-таблиці, але могли
    # бути створені не через Telegram-бота.
    sync_google_sheet_registrations_to_local()

    allowed = _web_allowed_group_ids(current)
    if not allowed:
        return {"registrations": []}

    if group_id is not None and group_id not in allowed:
        raise HTTPException(status_code=403, detail="Немає доступу до цього ресторану")

    query = """
        SELECT id, group_id, group_title, requester_name, child_name,
               phone_number, event_date, event_title,
               COALESCE(attendance_status, 'yes') AS attendance_status
        FROM mk_registrations
        WHERE event_date=?
    """
    params = [date]

    if group_id is not None:
        query += " AND group_id=?"
        params.append(group_id)
    else:
        placeholders = ",".join("?" for _ in allowed)
        query += f" AND group_id IN ({placeholders})"
        params.extend(allowed)

    search = (q or "").strip()
    if search:
        query += """
            AND (
                COALESCE(requester_name,'') LIKE ? OR
                COALESCE(child_name,'') LIKE ? OR
                COALESCE(phone_number,'') LIKE ? OR
                COALESCE(group_title,'') LIKE ?
            )
        """
        like = f"%{search}%"
        params.extend([like, like, like, like])

    query += " ORDER BY group_title COLLATE NOCASE, id"

    with _web_db() as conn:
        rows = conn.execute(query, params).fetchall()

    registrations = []
    for row in rows:
        phone = row["phone_number"] or ""
        registrations.append({
            "id": row["id"],
            "group_id": row["group_id"],
            "restaurant": row["group_title"] or str(row["group_id"]),
            "guest_name": row["requester_name"] or "—",
            "child_name": row["child_name"] or "—",
            "children": _web_children_count(row["child_name"]),
            "phone": phone,
            "status": row["attendance_status"] if row["attendance_status"] in {"yes", "no"} else "yes",
        })
    return {"registrations": registrations}


@web_app.patch("/api/registrations/{registration_id}/status")
async def web_update_status(registration_id: int, request: Request):
    current = _web_current_user(request)
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Некоректний JSON")

    status = payload.get("status")
    if status not in {"yes", "no"}:
        raise HTTPException(status_code=400, detail="Статус має бути yes або no")

    with _web_db() as conn:
        row = conn.execute(
            "SELECT group_id FROM mk_registrations WHERE id=?",
            (registration_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Запис не знайдено")
        if row["group_id"] not in _web_allowed_group_ids(current):
            raise HTTPException(status_code=403, detail="Немає доступу до цього ресторану")

        conn.execute(
            "UPDATE mk_registrations SET attendance_status=? WHERE id=?",
            (status, registration_id),
        )
        conn.commit()

    return {"ok": True, "status": status}


def start_web_server():
    port = int(os.getenv("PORT", "8000"))
    thread = threading.Thread(
        target=lambda: uvicorn.run(
            web_app,
            host="0.0.0.0",
            port=port,
            log_level="info",
        ),
        daemon=True,
        name="telegram-mini-app-web",
    )
    thread.start()
    return thread


# ============================================================
# ЗАПУСК БОТА
# ============================================================

def main():
    if not TOKEN:
        raise RuntimeError("BOT_TOKEN не заданий у змінних середовища")

    init_db()
    init_access_schema()
    init_mk_schedule_schema()
    check_db()
    start_web_server()

    app = (
        Application.builder()
        .token(TOKEN)
        .post_init(post_init)
        .build()
    )
    app.add_error_handler(error_handler)

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("groups", groups))
    app.add_handler(CommandHandler("stats", stats))
    app.add_handler(CommandHandler("addgroup", addgroup))
    app.add_handler(CommandHandler("excel", excel))
    app.add_handler(CommandHandler("broadcast", broadcast))
    app.add_handler(CommandHandler("deletebroadcast", deletebroadcast))
    app.add_handler(CommandHandler("delete_id", delete_by_id))
    app.add_handler(CommandHandler("setmkresponse", set_mk_response))
    app.add_handler(CommandHandler("setmkconfirm", set_mk_confirmation))
    app.add_handler(CommandHandler("setmkschedule", set_mk_schedule))
    app.add_handler(CommandHandler("mklist", mk_list))
    app.add_handler(CommandHandler("mkexcel", mk_excel))
    app.add_handler(CommandHandler("cancelmk", cancel_mk_setting))
    app.add_handler(CommandHandler("addmanager", add_manager))
    app.add_handler(CommandHandler("delmanager", delete_manager))
    app.add_handler(CommandHandler("addadmin", add_admin_role))
    app.add_handler(CommandHandler("deladmin", delete_admin_role))
    app.add_handler(CommandHandler("roles", roles_list))
    app.add_handler(CommandHandler("setname", set_user_custom_name))
    app.add_handler(CommandHandler("admins", admins_list))

    app.add_handler(
        MessageHandler(filters.CONTACT & filters.ChatType.PRIVATE, receive_mk_contact),
        group=-2,
    )

    app.add_handler(
        MessageHandler(
            filters.ChatType.PRIVATE & filters.TEXT & ~filters.COMMAND,
            receive_access_text,
        ),
        group=-3,
    )

    # Графік перехоплюється раніше за інші приватні тексти.
    app.add_handler(
        MessageHandler(
            filters.ChatType.PRIVATE & filters.TEXT & ~filters.COMMAND,
            receive_mk_schedule_text,
        ),
        group=-1,
    )

    app.add_handler(
        MessageHandler(
            filters.ChatType.PRIVATE & filters.TEXT & ~filters.COMMAND,
            receive_mk_setting_text,
        ),
        group=0,
    )

    app.add_handler(
        MessageHandler(
            filters.ChatType.GROUPS & filters.TEXT & ~filters.COMMAND,
            mk_registration_trigger,
        ),
        group=0,
    )

    app.add_handler(
        MessageHandler(~filters.COMMAND & ~filters.FORWARDED, receive_broadcast),
        group=1,
    )

    app.add_handler(
        CallbackQueryHandler(role_group_button, pattern=r"^rolegrp:"),
        group=-3,
    )
    app.add_handler(
        CallbackQueryHandler(mk_schedule_group_button, pattern=r"^mkschedgrp:"),
        group=-2,
    )
    app.add_handler(
        CallbackQueryHandler(mk_date_button, pattern=r"^mkdate:"),
        group=-1,
    )
    app.add_handler(
        CallbackQueryHandler(mk_setting_buttons, pattern=r"^mkcfg:"),
        group=0,
    )
    app.add_handler(
        CallbackQueryHandler(
            broadcast_buttons, pattern=r"^broadcast_(?:send|cancel)$"
        ),
        group=1,
    )

    app.add_handler(
        ChatMemberHandler(chat_member_update, ChatMemberHandler.CHAT_MEMBER)
    )

    print("🤖 БОТ УСПІШНО ЗАПУЩЕНИЙ")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logging.info("Бота зупинено користувачем.")
